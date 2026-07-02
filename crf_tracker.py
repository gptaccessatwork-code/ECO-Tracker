"""Queue and sync CRF numbers into the AMAT SGP CRF Tracker workbook.

This follows the same offline-queue pattern used by the ECO tracker:
- Outlook writes the CRF into a staging workbook in Downloads.
- The sync step waits for the OneDrive workbook to be stable and unlocked.
- Once the target is available, all queued CRF rows are written at once.
- Successfully processed rows are then removed from the staging queue.
"""

from __future__ import annotations

import argparse
import os
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook, load_workbook

from workbook_sync_utils import (
    log_event,
    log_exception,
    save_workbook_with_retry,
    wait_for_workbook_ready,
)


DEFAULT_STAGING_FILE = os.getenv(
    "CRF_STAGING_FILE",
    r"C:\Users\kmageshkumar\Downloads\AMAT SGP CRF Tracker.xlsx",
)
DEFAULT_STAGING_SHEET = os.getenv("CRF_STAGING_SHEET", "Sheet1")
DEFAULT_TARGET_FILE = os.getenv(
    "CRF_TARGET_FILE",
    r"C:\Users\kmageshkumar\OneDrive - Ichor Systems\AMAT SGP CRF Tracker.xlsx",
)
DEFAULT_TARGET_SHEET = os.getenv("CRF_TARGET_SHEET", "AMAT SGP CRF Tracker")
DEFAULT_RETRY_DELAY_MINUTES = 15
DEFAULT_SETTLE_SECONDS = 30
SYNC_LOG_FILE = Path(__file__).with_name(".crf_tracker_sync.log")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Queue CRF numbers and sync them into the local OneDrive CRF tracker workbook."
    )
    parser.add_argument(
        "--staging-workbook-file",
        default=DEFAULT_STAGING_FILE,
        help="Local queue workbook used as the offline staging area.",
    )
    parser.add_argument(
        "--staging-sheet-name",
        default=DEFAULT_STAGING_SHEET,
        help="Worksheet name for the offline staging queue.",
    )
    parser.add_argument(
        "--target-workbook-file",
        default=DEFAULT_TARGET_FILE,
        help="Locally synced OneDrive workbook that mirrors SharePoint.",
    )
    parser.add_argument(
        "--target-sheet-name",
        default=DEFAULT_TARGET_SHEET,
        help="Worksheet name in the target workbook.",
    )
    parser.add_argument(
        "--crf-number",
        required=True,
        help="CRF number extracted from the Outlook subject line.",
    )
    parser.add_argument(
        "--received-time",
        default="",
        help="Optional received time in 'YYYY-MM-DD HH:MM:SS' format.",
    )
    parser.add_argument(
        "--retry-delay-minutes",
        type=int,
        default=DEFAULT_RETRY_DELAY_MINUTES,
        help="Minutes to wait when the target workbook is busy.",
    )
    parser.add_argument(
        "--settle-seconds",
        type=int,
        default=DEFAULT_SETTLE_SECONDS,
        help="Seconds to wait after the target workbook looks free.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Show what would be queued and synced without modifying workbooks.",
    )

    args = parser.parse_args()
    if args.retry_delay_minutes < 1:
        parser.error("--retry-delay-minutes must be at least 1")
    if args.settle_seconds < 0:
        parser.error("--settle-seconds must be zero or greater")
    return args


def parse_received_time(value: str) -> datetime:
    if not value.strip():
        return datetime.now()

    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(value.strip(), fmt)
        except ValueError:
            continue

    raise ValueError(
        "received-time must use 'YYYY-MM-DD HH:MM:SS' or 'YYYY-MM-DDTHH:MM:SS'."
    )


def normalize_crf_number(value: Any) -> str:
    return str(value).strip()


def get_or_create_worksheet(workbook: Workbook, sheet_name: str):
    if sheet_name in workbook.sheetnames:
        return workbook[sheet_name]

    worksheet = workbook.create_sheet(title=sheet_name)
    return worksheet


def ensure_headers(worksheet) -> None:
    headers = ["CRF Number", "Received Date"]
    for index, header in enumerate(headers, start=1):
        existing = worksheet.cell(row=1, column=index).value
        if existing is None or str(existing).strip() != header:
            worksheet.cell(row=1, column=index, value=header)


def is_row_empty(values: List[Any]) -> bool:
    for value in values:
        if value is None:
            continue
        if isinstance(value, str) and not value.strip():
            continue
        return False
    return True


def load_rows_from_worksheet(path: str, sheet_name: str) -> List[Dict[str, Any]]:
    workbook = load_workbook(path, data_only=True)
    try:
        sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []

        headers = ["" if cell is None else str(cell).strip() for cell in rows[0]]
        data_rows: List[Dict[str, Any]] = []

        for row in rows[1:]:
            if is_row_empty(list(row)):
                continue

            item: Dict[str, Any] = {}
            for header, value in zip(headers, row):
                if not header or value is None:
                    continue
                item[header] = value
            if item:
                data_rows.append(item)

        return data_rows
    finally:
        workbook.close()


def get_existing_crfs_from_sheet(worksheet) -> set[str]:
    existing: set[str] = set()
    for row_index in range(2, worksheet.max_row + 1):
        value = worksheet.cell(row=row_index, column=1).value
        if value is None:
            continue
        crf_number = normalize_crf_number(value)
        if crf_number:
            existing.add(crf_number)
    return existing


def append_rows_to_worksheet(worksheet, rows: List[List[Any]]) -> None:
    next_row = max(2, worksheet.max_row + 1)
    for row_values in rows:
        for offset, value in enumerate(row_values, start=1):
            worksheet.cell(row=next_row, column=offset, value=value)
        next_row += 1


def read_staging_rows(staging_file: str, staging_sheet: str) -> List[Dict[str, Any]]:
    if not Path(staging_file).exists():
        return []
    return load_rows_from_worksheet(staging_file, staging_sheet)


def write_row_to_staging_queue(
    staging_file: str,
    staging_sheet: str,
    crf_number: str,
    received_time: datetime,
    retry_delay_minutes: int,
    settle_seconds: int,
    dry_run: bool,
) -> int:
    staging_path = Path(staging_file)
    workbook_exists = staging_path.exists()
    if workbook_exists:
        workbook = load_workbook(staging_path)
    else:
        workbook = Workbook()

    try:
        worksheet = get_or_create_worksheet(workbook, staging_sheet)
        ensure_headers(worksheet)

        normalized_crf = normalize_crf_number(crf_number)
        existing_crfs = get_existing_crfs_from_sheet(worksheet)
        if normalized_crf in existing_crfs:
            log_event(
                f"CRF {normalized_crf} already exists in the staging queue; skipping.",
                SYNC_LOG_FILE,
            )
            return 0

        received_value = received_time.strftime("%d %b %Y %H:%M:%S")
        if dry_run:
            print(f"[DRY RUN] Staging row: [{normalized_crf!r}, {received_value!r}]")
            return 0

        next_row = max(2, worksheet.max_row + 1)
        worksheet.cell(row=next_row, column=1, value=normalized_crf)
        worksheet.cell(row=next_row, column=2, value=received_value)

        if workbook_exists:
            save_workbook_with_retry(
                workbook,
                staging_path,
                retry_delay_minutes,
                settle_seconds,
                SYNC_LOG_FILE,
            )
        else:
            staging_path.parent.mkdir(parents=True, exist_ok=True)
            workbook.save(staging_path)
            log_event(
                f"Created new CRF staging workbook at {staging_path}.",
                SYNC_LOG_FILE,
            )

        log_event(f"CRF {normalized_crf} staged for sync.", SYNC_LOG_FILE)
        return 1
    finally:
        workbook.close()


def convert_source_rows_to_target_values(rows: List[Dict[str, Any]]) -> List[List[Any]]:
    converted: List[List[Any]] = []
    for row in rows:
        crf_number = normalize_crf_number(row.get("CRF Number", ""))
        received_date = row.get("Received Date", "")
        if crf_number:
            converted.append([crf_number, received_date])
    return converted


def write_rows_to_target_with_retry(
    target_file: str,
    target_sheet: str,
    rows: List[List[Any]],
    retry_delay_minutes: int,
    settle_seconds: int,
) -> int:
    target_path = Path(target_file)
    workbook_exists = target_path.exists()

    if workbook_exists:
        wait_for_workbook_ready(
            target_path,
            retry_delay_minutes,
            settle_seconds,
            SYNC_LOG_FILE,
        )
        workbook = load_workbook(target_path)
    else:
        workbook = Workbook()

    try:
        worksheet = get_or_create_worksheet(workbook, target_sheet)
        ensure_headers(worksheet)

        existing_crfs = get_existing_crfs_from_sheet(worksheet)
        rows_to_write: List[List[Any]] = []
        for row in rows:
            crf_number = normalize_crf_number(row[0] if row else "")
            if not crf_number:
                continue
            if crf_number in existing_crfs:
                log_event(f"Skipping duplicate CRF in target workbook: {crf_number}", SYNC_LOG_FILE)
                continue
            existing_crfs.add(crf_number)
            rows_to_write.append(row)

        if not rows_to_write:
            log_event("No new CRF rows to append after duplicate check.", SYNC_LOG_FILE)
            return 0

        append_rows_to_worksheet(worksheet, rows_to_write)

        if workbook_exists:
            save_workbook_with_retry(
                workbook,
                target_path,
                retry_delay_minutes,
                settle_seconds,
                SYNC_LOG_FILE,
            )
        else:
            target_path.parent.mkdir(parents=True, exist_ok=True)
            workbook.save(target_path)
            log_event(f"Created new CRF target workbook at {target_path}.", SYNC_LOG_FILE)

        log_event(f"CRF target workbook updated with {len(rows_to_write)} row(s).", SYNC_LOG_FILE)
        return len(rows_to_write)
    finally:
        workbook.close()


def remove_rows_from_staging_by_crf(
    staging_file: str,
    staging_sheet: str,
    crf_numbers: set[str],
    retry_delay_minutes: int,
    settle_seconds: int,
) -> None:
    if not crf_numbers:
        return

    staging_path = Path(staging_file)
    if not staging_path.exists():
        return

    workbook = load_workbook(staging_path)
    try:
        sheet = workbook[staging_sheet] if staging_sheet in workbook.sheetnames else workbook.active
        for row_idx in range(sheet.max_row, 1, -1):
            key_value = sheet.cell(row=row_idx, column=1).value
            if key_value is None:
                continue
            if normalize_crf_number(key_value) in crf_numbers:
                sheet.delete_rows(row_idx, 1)

        save_workbook_with_retry(
            workbook,
            staging_path,
            retry_delay_minutes,
            settle_seconds,
            SYNC_LOG_FILE,
        )
    finally:
        workbook.close()


def run_sync(
    staging_file: str,
    staging_sheet: str,
    target_file: str,
    target_sheet: str,
    retry_delay_minutes: int,
    settle_seconds: int,
) -> int:
    source_rows = read_staging_rows(staging_file, staging_sheet)
    if not source_rows:
        log_event("No queued CRF rows found in the staging workbook.", SYNC_LOG_FILE)
        return 0

    converted_rows = convert_source_rows_to_target_values(source_rows)
    crf_numbers_to_clear = {
        normalize_crf_number(row.get("CRF Number", ""))
        for row in source_rows
        if normalize_crf_number(row.get("CRF Number", ""))
    }

    written = write_rows_to_target_with_retry(
        target_file=target_file,
        target_sheet=target_sheet,
        rows=converted_rows,
        retry_delay_minutes=retry_delay_minutes,
        settle_seconds=settle_seconds,
    )

    if written or crf_numbers_to_clear:
        remove_rows_from_staging_by_crf(
            staging_file,
            staging_sheet,
            crf_numbers_to_clear,
            retry_delay_minutes,
            settle_seconds,
        )
        log_event("Processed CRF rows removed from the staging queue.", SYNC_LOG_FILE)

    return written


def main() -> int:
    args = parse_args()
    received_time = parse_received_time(args.received_time)

    normalized_crf = normalize_crf_number(args.crf_number)
    log_event(f"Processing CRF queue update for {normalized_crf}.", SYNC_LOG_FILE)

    staged = write_row_to_staging_queue(
        staging_file=args.staging_workbook_file,
        staging_sheet=args.staging_sheet_name,
        crf_number=normalized_crf,
        received_time=received_time,
        retry_delay_minutes=args.retry_delay_minutes,
        settle_seconds=args.settle_seconds,
        dry_run=args.dry_run,
    )

    if args.dry_run:
        print("Dry run complete. No workbook changes were made.")
        return 0

    if staged == 0:
        return 0

    written = run_sync(
        staging_file=args.staging_workbook_file,
        staging_sheet=args.staging_sheet_name,
        target_file=args.target_workbook_file,
        target_sheet=args.target_sheet_name,
        retry_delay_minutes=args.retry_delay_minutes,
        settle_seconds=args.settle_seconds,
    )

    print(f"Queued 1 row and synced {written} row(s) to the target workbook.")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        log_exception("Unhandled CRF tracker error.", exc, SYNC_LOG_FILE)
        print(f"Error: {exc}", file=sys.stderr)
        raise SystemExit(1)
