from __future__ import annotations

import sys
import types
import unittest
from pathlib import Path
from typing import Any
from unittest.mock import patch

from lxml import etree
from openpyxl import Workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

# Unit tests never load credentials, so keep the test suite independent of the
# shared Scripts-folder credential helper used in production.
credentials_module = types.ModuleType("script_credentials")
credentials_module.get_agile_credentials = lambda: ("", "")
credentials_module.credentials_file = lambda start=None: PROJECT_ROOT / "script_credentials.json"
sys.modules.setdefault("script_credentials", credentials_module)

import agile_eco_dates as subject


def make_row(**fields: str) -> dict[str, Any]:
    elements = []
    for name, value in fields.items():
        element = etree.Element(name)
        element.text = value
        elements.append(element)
    return {"_value_1": elements}


class FakeAgileClient:
    def __init__(self, rows_by_table: dict[str, list[dict[str, Any]]]) -> None:
        self.rows_by_table = rows_by_table
        self.calls: list[tuple[str, str]] = []

    def load_table(
        self,
        eco_number: str,
        class_identifier: str,
        table_identifier: str,
    ) -> list[dict[str, Any]]:
        self.calls.append((class_identifier, table_identifier))
        return self.rows_by_table.get(table_identifier, [])


class TrackingNumberTests(unittest.TestCase):
    def test_not_applicable_variants_are_unusable(self) -> None:
        variants = (
            "",
            " ",
            "NA",
            "na",
            "N/A",
            "N.A.",
            "N-A",
            "N A",
            "#N/A",
            "None",
            "not applicable",
            "Not-Applicable",
        )

        for value in variants:
            with self.subTest(value=value):
                self.assertTrue(subject.is_unusable_tracking_number(value))

    def test_real_identifiers_remain_usable(self) -> None:
        for value in ("CA90662HB", "DSM26134", "NA123", "N/A-123"):
            with self.subTest(value=value):
                self.assertFalse(subject.is_unusable_tracking_number(value))

    def test_only_explicit_not_applicable_values_request_delta_cleanup(self) -> None:
        for value in ("NA", "n/a", "N.A.", "None", "Not Applicable"):
            with self.subTest(value=value):
                self.assertTrue(subject.is_not_applicable_tracking_number(value))

        self.assertFalse(subject.is_not_applicable_tracking_number(""))
        self.assertFalse(subject.is_not_applicable_tracking_number("CA90662HB"))

    def test_delta_cleanup_preserves_non_delta_cells(self) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet["C2"] = "N/A"
        worksheet["E2"] = 14
        worksheet["F2"] = 10
        worksheet["H2"] = 14
        worksheet["I2"] = 10
        worksheet["J2"] = "Slot has been cancelled by AMAT."

        subject.clear_row_delta_cells(worksheet, 2, ["E", "F", "H", "I"])

        self.assertEqual(worksheet["C2"].value, "N/A")
        for column in ("E", "F", "H", "I"):
            self.assertIsNone(worksheet[f"{column}2"].value)
        self.assertEqual(worksheet["J2"].value, "Slot has been cancelled by AMAT.")
        workbook.close()


class SelfTestTests(unittest.TestCase):
    def test_credentials_result_uses_shared_credentials_path(self) -> None:
        with (
            patch.object(
                subject,
                "load_agile_credentials",
                return_value=("test-user", "test-password"),
            ),
            patch.object(subject, "dispatch_outlook_application", return_value=object()),
        ):
            results = subject.collect_self_test_results(None, None, "", "", "")

        credentials_result = next(
            result for result in results if result.label == "Agile credentials file"
        )
        self.assertTrue(credentials_result.ok)
        self.assertIn("script_credentials.json", credentials_result.detail)


class AgileDateExtractionTests(unittest.TestCase):
    def test_cover_page_date_released_is_recognized(self) -> None:
        submitted, released = subject.extract_status_dates(
            [
                make_row(
                    status="45603RELEASEDReleased",
                    datereleased="2026-07-29T00:00:45.000Z",
                )
            ]
        )

        self.assertIsNone(submitted)
        self.assertEqual(released, "29 Jul 2026")

    def test_change_status_next_status_is_recognized(self) -> None:
        submitted, released = subject.extract_status_dates(
            [
                make_row(
                    status="45600CCBCCB",
                    nextstatus="45603RELEASEDReleased",
                    action="65CHANGE_STATUSChange Status",
                    localclienttime="2026-07-28T02:38:43.000Z",
                    details="Austin Change Orders.CCB=>Austin Change Orders.Released",
                )
            ]
        )

        self.assertIsNone(submitted)
        self.assertEqual(released, "28 Jul 2026")

    def test_partial_history_result_is_completed_from_cover_page(self) -> None:
        client = FakeAgileClient(
            {
                "History": [
                    make_row(
                        status="45594PENDINGPending",
                        nextstatus="45597SUBMITTEDSubmitted",
                        action="65CHANGE_STATUSChange Status",
                        localclienttime="2026-07-27T13:02:02.000Z",
                        details="Pending=>Submitted",
                    )
                ],
                "Cover Page": [
                    make_row(
                        status="45603RELEASEDReleased",
                        datereleased="2026-07-29T00:00:45.000Z",
                    )
                ],
            }
        )

        result = subject.fetch_eco_dates(
            client=client,
            eco_number="TEST-ECO",
            class_identifiers=["ECO"],
            table_identifiers=["History", "Cover Page", "Workflow"],
            inspect=False,
        )

        self.assertEqual(result.submitted_date, "28 Jul 2026")
        self.assertEqual(result.released_date, "29 Jul 2026")
        self.assertEqual(result.table_identifier, "History, Cover Page")
        self.assertEqual(
            client.calls,
            [("ECO", "History"), ("ECO", "Cover Page")],
        )


class ConditionalFormattingTests(unittest.TestCase):
    def test_empty_data_range_is_ignored(self) -> None:
        workbook = Workbook()
        worksheet = workbook.active

        subject.refresh_delta_conditional_formatting(
            worksheet=worksheet,
            start_row=2,
            end_row=1,
            submitted_column="D",
            released_column="G",
            submitted_delta_column="E",
            released_delta_column="H",
            delta_columns=["E", "F", "H", "I"],
        )

        self.assertEqual(len(worksheet.conditional_formatting), 0)
        workbook.close()

    def test_tracker_threshold_formulas_are_preserved(self) -> None:
        workbook = Workbook()
        worksheet = workbook.active

        subject.refresh_delta_conditional_formatting(
            worksheet=worksheet,
            start_row=2,
            end_row=10,
            submitted_column="D",
            released_column="G",
            submitted_delta_column="E",
            released_delta_column="H",
            delta_columns=["E", "F", "H", "I"],
        )

        formulas = [
            rule.formula[0]
            for rules in worksheet.conditional_formatting._cf_rules.values()
            for rule in rules
        ]
        self.assertEqual(
            formulas,
            [
                'AND(ISNUMBER($E2),$E2>=2,$D2="")',
                'AND(ISNUMBER($H2),$D2<>"",$G2="",TODAY()-$D2>=3)',
            ],
        )
        workbook.close()


if __name__ == "__main__":
    unittest.main()
