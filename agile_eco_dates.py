"""Fetch ECO submitted/released dates from Agile Table WSDL.

This script connects to Agile PLM through the Table SOAP service, loads the
workflow-related table for a given ECO, and looks for workflow statuses named
"Submitted" and "Released". For each matching status, it returns the
"local client time" if present.

Credentials are read from the shared JSON file in the Scripts root:
- script_credentials.json
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from html import escape
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional
from zoneinfo import ZoneInfo

import win32com.client
from lxml import etree
from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from requests import Session
from requests.auth import HTTPBasicAuth
from zeep import Client, Settings, helpers
from zeep.transports import Transport
from workbook_sync_utils import (
    is_local_conflict_error,
    log_event,
    log_exception,
    save_workbook_with_retry,
    wait_for_workbook_ready,
)


AGILE_WSDL = "http://pagapps1.ichorsystems.com:7001/CoreService/services/Table?wsdl"
REQUEST_TABLE_QNAME = "{http://xmlns.oracle.com/AgileObjects/Core/Table/V1}RequestTableType"
LOAD_TABLE_REQUEST_QNAME = "{http://xmlns.oracle.com/AgileObjects/Core/Table/V1}LoadTableRequestType"
PORTLAND_TZ = ZoneInfo("America/Los_Angeles")
SINGAPORE_TZ = ZoneInfo("Asia/Singapore")

DEFAULT_CLASS_CANDIDATES = [
    "Change",
    "ECO",
    "Change Orders",
]

DEFAULT_TABLE_CANDIDATES = [
    "History",
    "Cover Page",
    "Workflow",
    "Workflow Status",
]

STATUS_FIELD_CANDIDATES = [
    "status",
    "nextstatus",
    "workflowstatus",
    "currentstatus",
    "name",
]

TIME_FIELD_CANDIDATES = [
    "localclienttime",
    "clienttime",
    "time",
    "date",
    "modifieddate",
]

ACTION_FIELD_CANDIDATES = [
    "action",
]

DETAILS_FIELD_CANDIDATES = [
    "details",
]

RELEASED_DATE_FIELD_CANDIDATES = [
    "datereleased",
    "releaseddate",
]

UNUSABLE_TRACKING_NUMBERS = frozenset({"", "na", "none", "notapplicable"})

DEFAULT_WORKBOOK_FILE = os.getenv("TARGET_WORKBOOK_FILE")
DEFAULT_WORKSHEET_NAME = os.getenv("TARGET_WORKSHEET")
DEFAULT_CRF_WORKBOOK_FILE = os.getenv("CRF_TARGET_WORKBOOK_FILE")
DEFAULT_CRF_WORKSHEET_NAME = os.getenv("CRF_TARGET_WORKSHEET")
DEFAULT_ECO_COLUMN = "C"
DEFAULT_SUBMITTED_COLUMN = "D"
DEFAULT_SUBMITTED_DELTA_COLUMN = "E"
DEFAULT_SUBMITTED_DELTA_EXCL_WEEKEND_COLUMN = "F"
DEFAULT_RELEASED_COLUMN = "G"
DEFAULT_RELEASED_DELTA_COLUMN = "H"
DEFAULT_RELEASED_DELTA_EXCL_WEEKEND_COLUMN = "I"
DEFAULT_CRF_NUMBER_COLUMN = "A"
DEFAULT_CRF_RECEIVED_COLUMN = "B"
DEFAULT_CRF_SUBMITTED_COLUMN = "C"
DEFAULT_CRF_SUBMITTED_DELTA_COLUMN = "D"
DEFAULT_CRF_SUBMITTED_DELTA_EXCL_WEEKEND_COLUMN = "E"
DEFAULT_CRF_RELEASED_COLUMN = "F"
DEFAULT_CRF_RELEASED_DELTA_COLUMN = "G"
DEFAULT_CRF_RELEASED_DELTA_EXCL_WEEKEND_COLUMN = "H"
DEFAULT_REMINDER_TO = os.getenv("ECO_REMINDER_TO", "")
DEFAULT_REMINDER_CC = os.getenv("ECO_REMINDER_CC", "")
DEFAULT_REMINDER_BCC = os.getenv("ECO_REMINDER_BCC", "")
DEFAULT_CRF_REMINDER_TO = os.getenv("CRF_REMINDER_TO", "")
DEFAULT_CRF_REMINDER_CC = os.getenv("CRF_REMINDER_CC", "")
DEFAULT_CRF_REMINDER_BCC = os.getenv("CRF_REMINDER_BCC", "")
REMINDER_STATE_FILE = Path(__file__).with_name(".eco_reminder_state.json")
CRF_REMINDER_STATE_FILE = Path(__file__).with_name(".crf_reminder_state.json")
LOCAL_WORKBOOK_RETRY_DELAY_MINUTES = 15
LOCAL_WORKBOOK_SETTLE_SECONDS = 30
SYNC_LOG_FILE = Path(__file__).with_name(".agile_eco_dates_sync.log")
ALERT_FILL = PatternFill(fill_type="solid", start_color="FF5B5B", end_color="FF5B5B")
ALERT_FONT = Font(color="000000", bold=True)
CLEAR_FILL = PatternFill(fill_type=None)
CLEAR_FONT = Font(color="000000", bold=False)
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
DISPLAY_DATE_FORMAT = "DD MMM YYYY"
DATE_PARSE_FORMATS = (
    "%Y-%m-%dT%H:%M:%S.%fZ",
    "%d %b %Y %I:%M:%S %p",
    "%d %b %Y %H:%M:%S",
    "%d %b %Y %I:%M %p",
    "%d %b %Y %H:%M",
    "%m/%d/%Y %I:%M:%S %p",
    "%m/%d/%Y %H:%M:%S",
    "%m/%d/%Y %I:%M %p",
    "%m/%d/%Y %H:%M",
    "%Y-%m-%dT%H:%M:%S",
    "%Y-%m-%d %H:%M:%S",
)

SCRIPT_ROOT = next(
    (parent for parent in Path(__file__).resolve().parents if (parent / "script_credentials.py").exists()),
    Path(__file__).resolve().parent,
)
if str(SCRIPT_ROOT) not in sys.path:
    sys.path.insert(0, str(SCRIPT_ROOT))

from script_credentials import credentials_file, get_agile_credentials
SHEET_DATE_FORMATS = ("%d %b %Y", "%d %B %Y", "%m/%d/%Y", "%Y-%m-%d", "%d/%m/%Y")
EMAIL_CELL_STYLE = (
    "border:1px solid #808080;padding:6px 10px;"
    "text-align:center;vertical-align:middle;"
)
EMAIL_HEADER_STYLE = (
    f"{EMAIL_CELL_STYLE}background-color:#d9e2f3;font-weight:bold;"
)


@dataclass(frozen=True)
class WorkbookColumns:
    spec_award: str
    eco: str
    submitted: str
    submitted_delta: str
    submitted_delta_excl_weekend: str
    released: str
    released_delta: str
    released_delta_excl_weekend: str


@dataclass(frozen=True)
class CrfWorkbookColumns:
    crf_number: str
    received_date: str
    submitted: str
    submitted_delta: str
    submitted_delta_excl_weekend: str
    released: str
    released_delta: str
    released_delta_excl_weekend: str


@dataclass
class EcoDates:
    submitted_date: Optional[str]
    released_date: Optional[str]
    class_identifier: str
    table_identifier: str


@dataclass
class ReminderItem:
    row_number: int
    system_number: str
    spec_award_date: str
    eco_number: str
    submitted_date: str
    released_date: str
    reason: str


@dataclass
class CrfReminderItem:
    row_number: int
    crf_number: str
    received_date: str
    submitted_date: str
    released_date: str
    reason: str


@dataclass
class SelfTestResult:
    label: str
    ok: bool
    detail: str


class AgileEcoClient:
    def __init__(self, username: str, password: str) -> None:
        session = Session()
        session.trust_env = False
        session.auth = HTTPBasicAuth(username, password)
        transport = Transport(session=session)
        settings = Settings(strict=False, xml_huge_tree=True)

        self.client = Client(AGILE_WSDL, transport=transport, settings=settings)
        self.RequestTableType = self.client.get_type(REQUEST_TABLE_QNAME)
        self.LoadTableRequestType = self.client.get_type(LOAD_TABLE_REQUEST_QNAME)

    def load_table(
        self,
        eco_number: str,
        class_identifier: str,
        table_identifier: str,
    ) -> List[Dict[str, Any]]:
        table_request = self.RequestTableType(
            classIdentifier=class_identifier,
            objectNumber=eco_number,
            tableIdentifier=table_identifier,
        )
        load_request = self.LoadTableRequestType(tableRequest=[table_request])
        response = self.client.service.loadTable(load_request)
        payload = helpers.serialize_object(response)
        table_contents = payload.get("tableContents") or []
        if not table_contents:
            return []
        return table_contents[0].get("row", []) or []


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Fetch Submitted/Released dates for ECOs from Agile."
    )
    parser.add_argument(
        "eco_number",
        nargs="?",
        help="Single ECO number to inspect. Omit this to update a workbook.",
    )
    parser.add_argument(
        "--crf-number",
        dest="crf_number",
        help="Single CRF number to inspect. Omit this to update the CRF workbook.",
    )
    parser.add_argument(
        "--workbook-file",
        default=DEFAULT_WORKBOOK_FILE,
        help="Local synced workbook file to update in place.",
    )
    parser.add_argument(
        "--crf-workbook-file",
        default=DEFAULT_CRF_WORKBOOK_FILE,
        help="Local synced CRF workbook file to update in place.",
    )
    parser.add_argument(
        "--worksheet",
        default=DEFAULT_WORKSHEET_NAME,
        help="Worksheet name to update. Defaults to the active sheet.",
    )
    parser.add_argument(
        "--crf-worksheet",
        default=DEFAULT_CRF_WORKSHEET_NAME,
        help="Worksheet name for the CRF workbook. Defaults to the active sheet.",
    )
    parser.add_argument(
        "--eco-column",
        default=DEFAULT_ECO_COLUMN,
        help="Column containing ECO numbers. Default: C",
    )
    parser.add_argument(
        "--submitted-column",
        default=DEFAULT_SUBMITTED_COLUMN,
        help="Column to write Submitted Date. Default: D",
    )
    parser.add_argument(
        "--released-column",
        default=DEFAULT_RELEASED_COLUMN,
        help="Column to write Released Date. Default: G",
    )
    parser.add_argument(
        "--spec-award-column",
        default="B",
        help="Column containing Spec Award Date. Default: B",
    )
    parser.add_argument(
        "--submitted-delta-column",
        default=DEFAULT_SUBMITTED_DELTA_COLUMN,
        help="Column to write Delta (Spec to Submitted). Default: E",
    )
    parser.add_argument(
        "--submitted-delta-excl-weekend-column",
        default=DEFAULT_SUBMITTED_DELTA_EXCL_WEEKEND_COLUMN,
        help="Column to write Delta (excl weekend). Default: F",
    )
    parser.add_argument(
        "--released-delta-column",
        default=DEFAULT_RELEASED_DELTA_COLUMN,
        help="Column to write Delta (Spec to Released). Default: H",
    )
    parser.add_argument(
        "--released-delta-excl-weekend-column",
        default=DEFAULT_RELEASED_DELTA_EXCL_WEEKEND_COLUMN,
        help="Column to write Delta (excl weekend)2. Default: I",
    )
    parser.add_argument(
        "--crf-number-column",
        default=DEFAULT_CRF_NUMBER_COLUMN,
        help="Column containing CRF numbers. Default: A",
    )
    parser.add_argument(
        "--crf-received-column",
        default=DEFAULT_CRF_RECEIVED_COLUMN,
        help="Column containing CRF received dates. Default: B",
    )
    parser.add_argument(
        "--crf-submitted-column",
        default=DEFAULT_CRF_SUBMITTED_COLUMN,
        help="Column to write Submitted Date in the CRF tracker. Default: C",
    )
    parser.add_argument(
        "--crf-submitted-delta-column",
        default=DEFAULT_CRF_SUBMITTED_DELTA_COLUMN,
        help="Column to write Delta (Received to Submitted) in the CRF tracker. Default: D",
    )
    parser.add_argument(
        "--crf-submitted-delta-excl-weekend-column",
        default=DEFAULT_CRF_SUBMITTED_DELTA_EXCL_WEEKEND_COLUMN,
        help="Column to write Delta (excl weekend) in the CRF tracker. Default: E",
    )
    parser.add_argument(
        "--crf-released-column",
        default=DEFAULT_CRF_RELEASED_COLUMN,
        help="Column to write Released Date in the CRF tracker. Default: F",
    )
    parser.add_argument(
        "--crf-released-delta-column",
        default=DEFAULT_CRF_RELEASED_DELTA_COLUMN,
        help="Column to write Delta (Received to Released) in the CRF tracker. Default: G",
    )
    parser.add_argument(
        "--crf-released-delta-excl-weekend-column",
        default=DEFAULT_CRF_RELEASED_DELTA_EXCL_WEEKEND_COLUMN,
        help="Column to write Delta (excl weekend) in the CRF tracker. Default: H",
    )
    parser.add_argument(
        "--start-row",
        type=int,
        default=2,
        help="First data row to process. Default: 2",
    )
    parser.add_argument(
        "--crf-start-row",
        type=int,
        default=2,
        help="First CRF data row to process. Default: 2",
    )
    parser.add_argument(
        "--class-identifier",
        action="append",
        dest="class_identifiers",
        help="Override Agile class identifier. Can be used multiple times.",
    )
    parser.add_argument(
        "--crf-class-identifier",
        action="append",
        dest="crf_class_identifiers",
        help="Override Agile class identifier for CRF. Can be used multiple times.",
    )
    parser.add_argument(
        "--table-identifier",
        action="append",
        dest="table_identifiers",
        help="Override Agile table identifier. Can be used multiple times.",
    )
    parser.add_argument(
        "--crf-table-identifier",
        action="append",
        dest="crf_table_identifiers",
        help="Override Agile table identifier for CRF. Can be used multiple times.",
    )
    parser.add_argument(
        "--inspect",
        action="store_true",
        help="Print discovered row fields to help tune class/table identifiers.",
    )
    parser.add_argument(
        "--crf-inspect",
        action="store_true",
        help="Print discovered row fields while inspecting CRF Agile data.",
    )
    parser.add_argument(
        "--self-test",
        action="store_true",
        help="Run setup checks without updating the workbook or sending email.",
    )
    parser.add_argument(
        "--crf-self-test",
        action="store_true",
        help="Run setup checks for the CRF workbook without updating it or sending email.",
    )
    parser.add_argument(
        "--preview-reminder",
        action="store_true",
        help="Open the reminder email as an Outlook draft instead of sending it.",
    )
    parser.add_argument("--reminder-to", default=DEFAULT_REMINDER_TO, help="Semicolon-separated reminder recipients.")
    parser.add_argument(
        "--reminder-cc",
        default=DEFAULT_REMINDER_CC,
        help="Semicolon-separated reminder CC recipients.",
    )
    parser.add_argument(
        "--reminder-bcc",
        default=DEFAULT_REMINDER_BCC,
        help="Semicolon-separated reminder BCC recipients.",
    )
    parser.add_argument(
        "--crf-preview-reminder",
        action="store_true",
        help="Open the CRF reminder email as an Outlook draft instead of sending it.",
    )
    parser.add_argument(
        "--crf-reminder-to",
        default=DEFAULT_CRF_REMINDER_TO,
        help="Semicolon-separated CRF reminder recipients.",
    )
    parser.add_argument(
        "--crf-reminder-cc",
        default=DEFAULT_CRF_REMINDER_CC,
        help="Semicolon-separated CRF reminder CC recipients.",
    )
    parser.add_argument(
        "--crf-reminder-bcc",
        default=DEFAULT_CRF_REMINDER_BCC,
        help="Semicolon-separated CRF reminder BCC recipients.",
    )
    return parser.parse_args()


def normalize_key(name: str) -> str:
    return "".join(ch.lower() for ch in name if ch.isalnum())


def is_unusable_tracking_number(value: str) -> bool:
    return normalize_key(value) in UNUSABLE_TRACKING_NUMBERS


def collapse_text(value: str) -> str:
    return " ".join(value.replace("\xa0", " ").split())


def extract_row_fields(row: Dict[str, Any]) -> Dict[str, str]:
    fields: Dict[str, str] = {}
    for elem in row.get("_value_1", []) or []:
        tag_name = normalize_key(etree.QName(elem.tag).localname)
        text_value = collapse_text("".join(elem.itertext()).strip())
        if text_value:
            fields[tag_name] = text_value
    return fields


def find_first_value(fields: Dict[str, str], candidates: Iterable[str]) -> Optional[str]:
    for candidate in candidates:
        key = normalize_key(candidate)
        if key in fields and fields[key]:
            return fields[key]
    return None


def parse_datetime(value: str) -> Optional[datetime]:
    for fmt in DATE_PARSE_FORMATS:
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    return None


def convert_agile_time_to_singapore(value: str) -> Optional[datetime]:
    parsed = parse_datetime(value)
    if parsed is None:
        return None

    # Agile's "local client time" is business-local Portland time for this workflow,
    # even when the value is serialized with a trailing "Z".
    if parsed.tzinfo is not None:
        parsed = parsed.replace(tzinfo=None)

    portland_time = parsed.replace(tzinfo=PORTLAND_TZ)
    return portland_time.astimezone(SINGAPORE_TZ)


def contains_status_name(value: str, status_name: str) -> bool:
    normalized_value = value.strip().lower()
    normalized_status = status_name.strip().lower()
    return normalized_status in normalized_value


def format_output_date(value: str) -> str:
    parsed = convert_agile_time_to_singapore(value)
    if parsed is None:
        return value
    return parsed.strftime("%d %b %Y")


def parse_sheet_date(value: Any) -> Optional[date]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text_value = str(value).strip()
    if not text_value:
        return None
    for fmt in SHEET_DATE_FORMATS:
        try:
            return datetime.strptime(text_value, fmt).date()
        except ValueError:
            continue
    return None


def calculate_delta_days(start_date: Optional[date], end_date: Optional[date]) -> Optional[int]:
    if not start_date or not end_date:
        return None
    return abs((end_date - start_date).days)


def count_weekend_days_between(start_date: date, end_date: date) -> int:
    if end_date == start_date:
        return 0
    if end_date < start_date:
        start_date, end_date = end_date, start_date
    weekend_days = 0
    current_date = start_date + timedelta(days=1)
    while current_date <= end_date:
        if current_date.weekday() >= 5:
            weekend_days += 1
        current_date += timedelta(days=1)
    return weekend_days


def calculate_delta_excluding_weekends(
    start_date: Optional[date],
    end_date: Optional[date],
) -> Optional[int]:
    if not start_date or not end_date:
        return None
    raw_delta = abs((end_date - start_date).days)
    weekend_days = count_weekend_days_between(start_date, end_date)
    return raw_delta - weekend_days


def singapore_today() -> date:
    return datetime.now(SINGAPORE_TZ).date()


def clear_cell_style(cell) -> None:
    cell.fill = CLEAR_FILL
    cell.font = CLEAR_FONT
    cell.alignment = CENTER_ALIGNMENT


def write_date_cell(cell, value: Optional[str]) -> None:
    parsed = parse_sheet_date(value)
    if parsed is None:
        cell.value = ""
    else:
        cell.value = parsed
        cell.number_format = DISPLAY_DATE_FORMAT
    cell.alignment = CENTER_ALIGNMENT


def apply_conditional_formatting(
    worksheet,
    range_ref: str,
    formula: str,
) -> None:
    worksheet.conditional_formatting.add(
        range_ref,
        FormulaRule(formula=[formula], fill=ALERT_FILL, font=ALERT_FONT),
    )


def clear_existing_conditional_formatting_for_columns(
    worksheet,
    start_row: int,
    end_row: int,
    columns: List[str],
) -> None:
    target_ranges = {f"{col}{start_row}:{col}{end_row}" for col in columns}
    cf_rules = getattr(worksheet.conditional_formatting, "_cf_rules", None)
    if cf_rules is None:
        return

    to_remove = []
    for key in list(cf_rules.keys()):
        sqref_text = str(getattr(key, "sqref", key))
        if sqref_text in target_ranges:
            to_remove.append(key)

    for key in to_remove:
        del cf_rules[key]


def refresh_delta_conditional_formatting(
    worksheet,
    start_row: int,
    end_row: int,
    submitted_column: str,
    released_column: str,
    submitted_delta_column: str,
    released_delta_column: str,
    delta_columns: List[str],
) -> None:
    if end_row < start_row:
        return

    clear_existing_conditional_formatting_for_columns(
        worksheet,
        start_row,
        end_row,
        delta_columns,
    )
    apply_conditional_formatting(
        worksheet,
        f"{submitted_delta_column}{start_row}:{submitted_delta_column}{end_row}",
        (
            f"AND(ISNUMBER(${submitted_delta_column}{start_row}),"
            f"${submitted_delta_column}{start_row}>=2,"
            f'${submitted_column}{start_row}="")'
        ),
    )
    apply_conditional_formatting(
        worksheet,
        f"{released_delta_column}{start_row}:{released_delta_column}{end_row}",
        (
            f'AND(${submitted_column}{start_row}<>"",'
            f'${released_column}{start_row}="",'
            f"TODAY()-${submitted_column}{start_row}>=3)"
        ),
    )


def load_json_state(state_file: Path) -> Dict[str, bool]:
    if not state_file.exists():
        return {}
    try:
        return json.loads(state_file.read_text(encoding="utf-8"))
    except Exception:
        return {}


def save_json_state(state_file: Path, state: Dict[str, bool]) -> None:
    state_file.write_text(
        json.dumps(state, indent=2, sort_keys=True),
        encoding="utf-8",
    )


def load_reminder_state() -> Dict[str, bool]:
    return load_json_state(REMINDER_STATE_FILE)


def save_reminder_state(state: Dict[str, bool]) -> None:
    save_json_state(REMINDER_STATE_FILE, state)


def load_crf_reminder_state() -> Dict[str, bool]:
    return load_json_state(CRF_REMINDER_STATE_FILE)


def save_crf_reminder_state(state: Dict[str, bool]) -> None:
    save_json_state(CRF_REMINDER_STATE_FILE, state)


def reminder_key(reminder: ReminderItem) -> str:
    return "|".join(
        [
            singapore_today().isoformat(),
            str(reminder.row_number),
            reminder.system_number,
            reminder.eco_number,
            reminder.reason,
        ]
    )


def load_agile_credentials() -> tuple[str, str]:
    return get_agile_credentials()


def create_agile_client_from_file() -> AgileEcoClient:
    agile_user, agile_pass = load_agile_credentials()
    return AgileEcoClient(agile_user, agile_pass)


def dispatch_outlook_application():
    return win32com.client.Dispatch("Outlook.Application")


def build_email_cell(value: str, style: str = EMAIL_CELL_STYLE) -> str:
    return f'<td style="{style}">{escape(value)}</td>'


def build_reminder_rows_html(reminders: List[ReminderItem]) -> str:
    rows_html: List[str] = []
    for item in reminders:
        eco_text = item.eco_number if item.eco_number else "Missing"
        submitted_text = item.submitted_date if item.submitted_date else "Pending"
        released_text = item.released_date if item.released_date else "Pending"
        cells = [
            item.system_number,
            item.spec_award_date,
            eco_text,
            submitted_text,
            released_text,
            item.reason,
        ]
        rows_html.append("<tr>" + "".join(build_email_cell(cell) for cell in cells) + "</tr>")
    return "".join(rows_html)


def send_reminder_email(
    reminders: List[ReminderItem],
    to_recipients: str,
    cc_recipients: str,
    bcc_recipients: str,
    preview_only: bool = False,
) -> bool:
    if not reminders:
        print("No reminder email created: no overdue rows matched the follow-up rules.")
        return False

    if not to_recipients.strip():
        print("No reminder email created: reminder recipients are empty. Use --reminder-to or set ECO_REMINDER_TO.")
        return False

    outlook = dispatch_outlook_application()
    mail = outlook.CreateItem(0)
    mail.To = to_recipients
    mail.CC = cc_recipients
    mail.BCC = bcc_recipients
    mail.Subject = f"ECO Tracker Follow-Up - {singapore_today().strftime('%d %b %Y')}"
    mail.HTMLBody = f"""
<html>
  <body style="font-family:Calibri,Arial,sans-serif;font-size:11pt;color:#1f1f1f;">
    <p>Hi team,</p>
    <p>The following items need follow-up:</p>
    <table style="border-collapse:collapse;border:1px solid #808080;">
      <tr style="background-color:#d9e2f3;font-weight:bold;">
        {build_email_cell("System Number", EMAIL_HEADER_STYLE)}
        {build_email_cell("Spec Award Date", EMAIL_HEADER_STYLE)}
        {build_email_cell("ECO Number", EMAIL_HEADER_STYLE)}
        {build_email_cell("Submitted Date", EMAIL_HEADER_STYLE)}
        {build_email_cell("Released Date", EMAIL_HEADER_STYLE)}
        {build_email_cell("Action Needed", EMAIL_HEADER_STYLE)}
      </tr>
      {build_reminder_rows_html(reminders)}
    </table>
    <p>Please follow up accordingly.</p>
  </body>
</html>
"""
    if preview_only:
        mail.Display()
        print(f"Reminder preview opened for: {to_recipients}")
    else:
        mail.Send()
        print(f"Reminder email sent to: {to_recipients}")
    return True


def crf_reminder_key(reminder: CrfReminderItem) -> str:
    return "|".join(
        [
            singapore_today().isoformat(),
            str(reminder.row_number),
            reminder.crf_number,
            reminder.reason,
        ]
    )


def build_crf_reminder_rows_html(reminders: List[CrfReminderItem]) -> str:
    rows_html: List[str] = []
    for item in reminders:
        submitted_text = item.submitted_date if item.submitted_date else "Pending"
        released_text = item.released_date if item.released_date else "Pending"
        cells = [
            item.crf_number,
            item.received_date,
            submitted_text,
            released_text,
            item.reason,
        ]
        rows_html.append("<tr>" + "".join(build_email_cell(cell) for cell in cells) + "</tr>")
    return "".join(rows_html)


def send_crf_reminder_email(
    reminders: List[CrfReminderItem],
    to_recipients: str,
    cc_recipients: str,
    bcc_recipients: str,
    preview_only: bool = False,
) -> bool:
    if not reminders:
        print("No CRF reminder email created: no overdue rows matched the follow-up rules.")
        return False

    if not to_recipients.strip():
        print(
            "No CRF reminder email created: reminder recipients are empty. "
            "Use --crf-reminder-to or set CRF_REMINDER_TO."
        )
        return False

    outlook = dispatch_outlook_application()
    mail = outlook.CreateItem(0)
    mail.To = to_recipients
    mail.CC = cc_recipients
    mail.BCC = bcc_recipients
    mail.Subject = f"CRF Tracker Follow-Up - {singapore_today().strftime('%d %b %Y')}"
    mail.HTMLBody = f"""
<html>
  <body style="font-family:Calibri,Arial,sans-serif;font-size:11pt;color:#1f1f1f;">
    <p>Hi team,</p>
    <p>The following CRF items need follow-up:</p>
    <table style="border-collapse:collapse;border:1px solid #808080;">
      <tr style="background-color:#d9e2f3;font-weight:bold;">
        {build_email_cell("CRF Number", EMAIL_HEADER_STYLE)}
        {build_email_cell("Received Date", EMAIL_HEADER_STYLE)}
        {build_email_cell("Submitted Date", EMAIL_HEADER_STYLE)}
        {build_email_cell("Released Date", EMAIL_HEADER_STYLE)}
        {build_email_cell("Action Needed", EMAIL_HEADER_STYLE)}
      </tr>
      {build_crf_reminder_rows_html(reminders)}
    </table>
    <p>Please follow up accordingly.</p>
  </body>
</html>
"""
    if preview_only:
        mail.Display()
        print(f"CRF reminder preview opened for: {to_recipients}")
    else:
        mail.Send()
        print(f"CRF reminder email sent to: {to_recipients}")
    return True


def collect_self_test_results(
    workbook_file: Optional[str],
    worksheet_name: Optional[str],
    reminder_to: str,
    reminder_cc: str,
    reminder_bcc: str,
) -> List[SelfTestResult]:
    results: List[SelfTestResult] = []

    try:
        agile_credentials_file = credentials_file(Path(__file__))
        agile_user, agile_pass = load_agile_credentials()
        results.append(
            SelfTestResult(
                label="Agile credentials file",
                ok=True,
                detail=f"Loaded from {agile_credentials_file}.",
            )
        )
        results.append(
            SelfTestResult(
                label="AGILE_USER",
                ok=bool(agile_user),
                detail="Found in JSON." if agile_user else "Missing from JSON.",
            )
        )
        results.append(
            SelfTestResult(
                label="AGILE_PASS",
                ok=bool(agile_pass),
                detail="Found in JSON." if agile_pass else "Missing from JSON.",
            )
        )
    except Exception as exc:
        results.append(
            SelfTestResult(
                label="Agile credentials file",
                ok=False,
                detail=str(exc),
            )
        )

    try:
        dispatch_outlook_application()
        results.append(
            SelfTestResult(
                label="Outlook",
                ok=True,
                detail="Outlook COM is available.",
            )
        )
    except Exception as exc:
        results.append(
            SelfTestResult(
                label="Outlook",
                ok=False,
                detail=f"Outlook COM unavailable: {exc}",
            )
        )

    recipient_summary = []
    if reminder_to.strip():
        recipient_summary.append(f"To={reminder_to}")
    if reminder_cc.strip():
        recipient_summary.append(f"Cc={reminder_cc}")
    if reminder_bcc.strip():
        recipient_summary.append(f"Bcc={reminder_bcc}")
    results.append(
        SelfTestResult(
            label="Reminder Recipients",
            ok=bool(reminder_to.strip()),
            detail="; ".join(recipient_summary)
            if recipient_summary
            else "No reminder recipients configured.",
        )
    )

    if workbook_file:
        workbook_path = Path(workbook_file)
        results.append(
            SelfTestResult(
                label="Workbook File",
                ok=workbook_path.exists(),
                detail=str(workbook_path) if workbook_path.exists() else f"File not found: {workbook_file}",
            )
        )
        if workbook_path.exists():
            workbook = load_workbook(workbook_path)
            try:
                if worksheet_name:
                    ok = worksheet_name in workbook.sheetnames
                    detail = (
                        f"Worksheet '{worksheet_name}' found."
                        if ok
                        else f"Worksheet '{worksheet_name}' not found. Available: {', '.join(workbook.sheetnames)}"
                    )
                    results.append(SelfTestResult(label="Worksheet", ok=ok, detail=detail))
                else:
                    results.append(
                        SelfTestResult(
                            label="Worksheet",
                            ok=True,
                            detail=f"Using active worksheet '{workbook.active.title}'.",
                        )
                    )
            finally:
                workbook.close()
    else:
        results.append(
            SelfTestResult(
                label="Workbook File",
                ok=False,
                detail="Not provided. Use --workbook-file or TARGET_WORKBOOK_FILE.",
            )
        )

    return results


def run_self_test(
    workbook_file: Optional[str],
    worksheet_name: Optional[str],
    reminder_to: str,
    reminder_cc: str,
    reminder_bcc: str,
) -> int:
    results = collect_self_test_results(
        workbook_file=workbook_file,
        worksheet_name=worksheet_name,
        reminder_to=reminder_to,
        reminder_cc=reminder_cc,
        reminder_bcc=reminder_bcc,
    )
    has_failure = False
    for result in results:
        status = "PASS" if result.ok else "FAIL"
        print(f"{status}: {result.label} - {result.detail}")
        if not result.ok:
            has_failure = True
    return 1 if has_failure else 0


def build_follow_up_reason(
    spec_award_date: Optional[date],
    eco_number: str,
    submitted_date: Optional[date],
    released_date: Optional[date],
    today_sg: date,
) -> str:
    if not spec_award_date:
        return ""

    days_since_spec = (today_sg - spec_award_date).days
    if not eco_number and days_since_spec >= 2:
        return "No ECO number after 2 days from Spec Award date"

    if not submitted_date and days_since_spec >= 2:
        return "ECO not submitted within 2 days from Spec Award date"

    if submitted_date and not released_date:
        days_since_submitted = (today_sg - submitted_date).days
        if days_since_submitted >= 3:
            return "ECO not released after 3 days from Submitted date"

    return ""


def build_crf_follow_up_reason(
    received_date: Optional[date],
    crf_number: str,
    submitted_date: Optional[date],
    released_date: Optional[date],
    today_sg: date,
) -> str:
    if not received_date:
        return ""

    days_since_received = (today_sg - received_date).days
    if not crf_number and days_since_received >= 2:
        return "No CRF number after 2 days from Received date"

    if not submitted_date and days_since_received >= 2:
        return "CRF not submitted within 2 days from Received date"

    if submitted_date and not released_date:
        days_since_submitted = (today_sg - submitted_date).days
        if days_since_submitted >= 3:
            return "CRF not released after 3 days from Submitted date"

    return ""


def extract_status_dates(
    rows: List[Dict[str, Any]],
    inspect: bool = False,
) -> tuple[Optional[str], Optional[str]]:
    submitted_matches: List[tuple[int, str]] = []
    released_matches: List[tuple[int, str]] = []

    for index, row in enumerate(rows, start=1):
        fields = extract_row_fields(row)
        if inspect:
            print(f"Row {index} fields: {fields}")

        explicit_released_date = find_first_value(
            fields, RELEASED_DATE_FIELD_CANDIDATES
        )
        if explicit_released_date:
            released_matches.append((2, explicit_released_date))

        action_value = find_first_value(fields, ACTION_FIELD_CANDIDATES) or ""
        time_value = find_first_value(fields, TIME_FIELD_CANDIDATES)
        if not time_value:
            continue

        status_candidates = []
        for candidate in STATUS_FIELD_CANDIDATES:
            field_key = normalize_key(candidate)
            if field_key in fields and fields[field_key]:
                status_candidates.append(fields[field_key])

        is_status_change = "changestatus" in normalize_key(action_value)
        if is_status_change:
            details_value = find_first_value(fields, DETAILS_FIELD_CANDIDATES)
            if details_value:
                status_candidates.append(details_value)

        if not status_candidates:
            continue

        row_weight = 1 if is_status_change else 0

        if any(contains_status_name(value, "submitted") for value in status_candidates):
            submitted_matches.append((row_weight, time_value))
        if any(contains_status_name(value, "released") for value in status_candidates):
            released_matches.append((row_weight, time_value))

    submitted_date = pick_best_date(submitted_matches)
    released_date = pick_best_date(released_matches)
    return submitted_date, released_date


def pick_best_date(values: List[tuple[int, str]]) -> Optional[str]:
    if not values:
        return None

    parsed_values = [(weight, parse_datetime(value), value) for weight, value in values]
    parsed_only = [pair for pair in parsed_values if pair[1] is not None]
    if parsed_only:
        parsed_only.sort(key=lambda pair: (-pair[0], pair[1]))
        return format_output_date(parsed_only[0][2])
    return values[0][1]


def fetch_eco_dates(
    client: AgileEcoClient,
    eco_number: str,
    class_identifiers: List[str],
    table_identifiers: List[str],
    inspect: bool,
) -> EcoDates:
    errors: List[str] = []
    submitted_date: Optional[str] = None
    released_date: Optional[str] = None
    matched_classes: List[str] = []
    matched_tables: List[str] = []
    first_nonempty_source: Optional[tuple[str, str]] = None

    for class_identifier in class_identifiers:
        for table_identifier in table_identifiers:
            try:
                rows = client.load_table(eco_number, class_identifier, table_identifier)
            except Exception as exc:
                errors.append(f"{class_identifier}/{table_identifier}: {exc}")
                continue

            if not rows:
                continue

            if first_nonempty_source is None:
                first_nonempty_source = (class_identifier, table_identifier)

            table_submitted_date, table_released_date = extract_status_dates(
                rows, inspect=inspect
            )
            found_date = False
            if submitted_date is None and table_submitted_date:
                submitted_date = table_submitted_date
                found_date = True
            if released_date is None and table_released_date:
                released_date = table_released_date
                found_date = True
            if found_date:
                if class_identifier not in matched_classes:
                    matched_classes.append(class_identifier)
                if table_identifier not in matched_tables:
                    matched_tables.append(table_identifier)

            if submitted_date and released_date:
                return EcoDates(
                    submitted_date=submitted_date,
                    released_date=released_date,
                    class_identifier=", ".join(matched_classes),
                    table_identifier=", ".join(matched_tables),
                )

    if submitted_date or released_date:
        return EcoDates(
            submitted_date=submitted_date,
            released_date=released_date,
            class_identifier=", ".join(matched_classes),
            table_identifier=", ".join(matched_tables),
        )

    if inspect and first_nonempty_source:
        return EcoDates(
            submitted_date=None,
            released_date=None,
            class_identifier=first_nonempty_source[0],
            table_identifier=first_nonempty_source[1],
        )

    error_text = "\n".join(errors[-5:]) if errors else "No matching workflow rows found."
    raise RuntimeError(
        "Could not find Submitted/Released workflow dates for "
        f"{eco_number}.\nTried classes={class_identifiers} tables={table_identifiers}\n{error_text}"
    )


def update_workbook_dates(
    workbook_file: str,
    worksheet_name: Optional[str],
    spec_award_column: str,
    eco_column: str,
    submitted_column: str,
    submitted_delta_column: str,
    submitted_delta_excl_weekend_column: str,
    released_column: str,
    released_delta_column: str,
    released_delta_excl_weekend_column: str,
    start_row: int,
    class_identifiers: List[str],
    table_identifiers: List[str],
    reminder_to: str,
    reminder_cc: str,
    reminder_bcc: str,
    preview_reminder: bool,
) -> int:
    workbook_path = Path(workbook_file)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook file not found: {workbook_file}")

    columns = WorkbookColumns(
        spec_award=spec_award_column.upper(),
        eco=eco_column.upper(),
        submitted=submitted_column.upper(),
        submitted_delta=submitted_delta_column.upper(),
        submitted_delta_excl_weekend=submitted_delta_excl_weekend_column.upper(),
        released=released_column.upper(),
        released_delta=released_delta_column.upper(),
        released_delta_excl_weekend=released_delta_excl_weekend_column.upper(),
    )

    while True:
        wait_for_workbook_ready(
            workbook_path,
            LOCAL_WORKBOOK_RETRY_DELAY_MINUTES,
            LOCAL_WORKBOOK_SETTLE_SECONDS,
            SYNC_LOG_FILE,
        )

        workbook = None
        try:
            updated_rows = 0
            skipped_released_rows = 0
            skipped_unusable_rows = 0
            eco_cache: Dict[str, EcoDates] = {}
            today_sg = singapore_today()
            reminders: List[ReminderItem] = []
            reminder_state = load_reminder_state()
            sent_reminder_keys: List[str] = []
            agile_client: Optional[AgileEcoClient] = None
            workbook = load_workbook(workbook_path)
            worksheet = workbook[worksheet_name] if worksheet_name else workbook.active

            for row_idx in range(start_row, worksheet.max_row + 1):
                eco_value = worksheet[f"{columns.eco}{row_idx}"].value
                existing_released_date = parse_sheet_date(
                    worksheet[f"{columns.released}{row_idx}"].value
                )
                spec_award_date = parse_sheet_date(
                    worksheet[f"{columns.spec_award}{row_idx}"].value
                )
                system_number_value = worksheet[f"A{row_idx}"].value
                system_number = (
                    "" if system_number_value is None else str(system_number_value).strip()
                )
                eco_number = "" if eco_value is None else str(eco_value).strip()

                if existing_released_date:
                    skipped_released_rows += 1
                    continue

                if is_unusable_tracking_number(eco_number):
                    skipped_unusable_rows += 1
                    continue

                if eco_number not in eco_cache:
                    try:
                        if agile_client is None:
                            agile_client = create_agile_client_from_file()
                        eco_cache[eco_number] = fetch_eco_dates(
                            client=agile_client,
                            eco_number=eco_number,
                            class_identifiers=class_identifiers,
                            table_identifiers=table_identifiers,
                            inspect=False,
                        )
                    except RuntimeError as exc:
                        log_event(
                            f"Row {row_idx}: ECO lookup failed for {eco_number}; "
                            f"treating as not submitted yet. {exc}",
                            SYNC_LOG_FILE,
                        )
                        eco_cache[eco_number] = EcoDates(
                            submitted_date=None,
                            released_date=None,
                            class_identifier="",
                            table_identifier="",
                        )
                result = eco_cache[eco_number]

                submitted_date_text = result.submitted_date if result else ""
                released_date_text = result.released_date if result else ""
                write_date_cell(worksheet[f"{columns.submitted}{row_idx}"], submitted_date_text)
                write_date_cell(worksheet[f"{columns.released}{row_idx}"], released_date_text)

                submitted_date = parse_sheet_date(submitted_date_text)
                released_date = parse_sheet_date(released_date_text)
                submitted_delta_end = submitted_date or today_sg
                released_delta_end = released_date or today_sg

                cell_updates = [
                    (
                        worksheet[f"{columns.submitted_delta}{row_idx}"],
                        calculate_delta_days(spec_award_date, submitted_delta_end),
                    ),
                    (
                        worksheet[f"{columns.submitted_delta_excl_weekend}{row_idx}"],
                        calculate_delta_excluding_weekends(spec_award_date, submitted_delta_end),
                    ),
                    (
                        worksheet[f"{columns.released_delta}{row_idx}"],
                        calculate_delta_days(spec_award_date, released_delta_end),
                    ),
                    (
                        worksheet[f"{columns.released_delta_excl_weekend}{row_idx}"],
                        calculate_delta_excluding_weekends(spec_award_date, released_delta_end),
                    ),
                ]
                for cell, value in cell_updates:
                    cell.value = value
                    clear_cell_style(cell)

                follow_up_reason = build_follow_up_reason(
                    spec_award_date=spec_award_date,
                    eco_number=eco_number,
                    submitted_date=submitted_date,
                    released_date=released_date,
                    today_sg=today_sg,
                )
                if follow_up_reason and spec_award_date:
                    reminder = ReminderItem(
                        row_number=row_idx,
                        system_number=system_number,
                        spec_award_date=spec_award_date.strftime("%d %b %Y"),
                        eco_number=eco_number,
                        submitted_date=submitted_date.strftime("%d %b %Y") if submitted_date else "",
                        released_date=released_date.strftime("%d %b %Y") if released_date else "",
                        reason=follow_up_reason,
                    )
                    key = reminder_key(reminder)
                    if not reminder_state.get(key):
                        reminders.append(reminder)
                        sent_reminder_keys.append(key)
                    else:
                        log_event(
                            f"Reminder already sent today for row {row_idx}: "
                            f"{system_number or '(blank system number)'} - {follow_up_reason}",
                            SYNC_LOG_FILE,
                        )

                updated_rows += 1
                log_event(
                    f"Updated row {row_idx}: ECO={eco_number} "
                    f"Submitted={submitted_date_text} Released={released_date_text}",
                    SYNC_LOG_FILE,
                )

            refresh_delta_conditional_formatting(
                worksheet,
                start_row,
                worksheet.max_row,
                columns.submitted,
                columns.released,
                columns.submitted_delta,
                columns.released_delta,
                [
                    columns.submitted_delta,
                    columns.released_delta,
                    columns.submitted_delta_excl_weekend,
                    columns.released_delta_excl_weekend,
                ],
            )

            save_workbook_with_retry(
                workbook,
                workbook_path,
                LOCAL_WORKBOOK_RETRY_DELAY_MINUTES,
                LOCAL_WORKBOOK_SETTLE_SECONDS,
                SYNC_LOG_FILE,
            )
            break
        except Exception as exc:
            if not is_local_conflict_error(exc):
                log_exception("Unhandled workbook update error.", exc, SYNC_LOG_FILE)
                raise
            log_exception("Workbook update conflict detected; will retry.", exc, SYNC_LOG_FILE)
            log_event(
                f"Workbook is still open or locked. Retrying in {LOCAL_WORKBOOK_RETRY_DELAY_MINUTES} minutes.",
                SYNC_LOG_FILE,
            )
            continue
        finally:
            if workbook is not None:
                workbook.close()

    reminder_sent = False
    if reminders:
        reminder_sent = send_reminder_email(
            reminders,
            reminder_to,
            reminder_cc,
            reminder_bcc,
            preview_only=preview_reminder,
        )
    if reminder_sent and not preview_reminder:
        for key in sent_reminder_keys:
            reminder_state[key] = True
        save_reminder_state(reminder_state)
    log_event(
        f"Workbook updated: {workbook_file}; rows updated={updated_rows}; "
        f"rows skipped (already released)={skipped_released_rows}; "
        f"rows skipped (blank/not applicable)={skipped_unusable_rows}",
        SYNC_LOG_FILE,
    )
    return updated_rows


def update_crf_workbook_dates(
    workbook_file: str,
    worksheet_name: Optional[str],
    crf_number_column: str,
    received_date_column: str,
    submitted_column: str,
    submitted_delta_column: str,
    submitted_delta_excl_weekend_column: str,
    released_column: str,
    released_delta_column: str,
    released_delta_excl_weekend_column: str,
    start_row: int,
    class_identifiers: List[str],
    table_identifiers: List[str],
    reminder_to: str,
    reminder_cc: str,
    reminder_bcc: str,
    preview_reminder: bool,
) -> int:
    workbook_path = Path(workbook_file)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook file not found: {workbook_file}")

    columns = CrfWorkbookColumns(
        crf_number=crf_number_column.upper(),
        received_date=received_date_column.upper(),
        submitted=submitted_column.upper(),
        submitted_delta=submitted_delta_column.upper(),
        submitted_delta_excl_weekend=submitted_delta_excl_weekend_column.upper(),
        released=released_column.upper(),
        released_delta=released_delta_column.upper(),
        released_delta_excl_weekend=released_delta_excl_weekend_column.upper(),
    )

    while True:
        wait_for_workbook_ready(
            workbook_path,
            LOCAL_WORKBOOK_RETRY_DELAY_MINUTES,
            LOCAL_WORKBOOK_SETTLE_SECONDS,
            SYNC_LOG_FILE,
        )

        workbook = None
        try:
            updated_rows = 0
            skipped_released_rows = 0
            skipped_unusable_rows = 0
            crf_cache: Dict[str, EcoDates] = {}
            today_sg = singapore_today()
            reminders: List[CrfReminderItem] = []
            reminder_state = load_crf_reminder_state()
            sent_reminder_keys: List[str] = []
            agile_client: Optional[AgileEcoClient] = None
            workbook = load_workbook(workbook_path)
            worksheet = workbook[worksheet_name] if worksheet_name else workbook.active

            for row_idx in range(start_row, worksheet.max_row + 1):
                crf_value = worksheet[f"{columns.crf_number}{row_idx}"].value
                existing_released_date = parse_sheet_date(
                    worksheet[f"{columns.released}{row_idx}"].value
                )
                received_date = parse_sheet_date(
                    worksheet[f"{columns.received_date}{row_idx}"].value
                )
                crf_number = "" if crf_value is None else str(crf_value).strip()

                if existing_released_date:
                    skipped_released_rows += 1
                    continue

                if is_unusable_tracking_number(crf_number):
                    skipped_unusable_rows += 1
                    continue

                if crf_number not in crf_cache:
                    try:
                        if agile_client is None:
                            agile_client = create_agile_client_from_file()
                        crf_cache[crf_number] = fetch_eco_dates(
                            client=agile_client,
                            eco_number=crf_number,
                            class_identifiers=class_identifiers,
                            table_identifiers=table_identifiers,
                            inspect=False,
                        )
                    except RuntimeError as exc:
                        log_event(
                            f"Row {row_idx}: CRF lookup failed for {crf_number}; "
                            f"treating as not submitted yet. {exc}",
                            SYNC_LOG_FILE,
                        )
                        crf_cache[crf_number] = EcoDates(
                            submitted_date=None,
                            released_date=None,
                            class_identifier="",
                            table_identifier="",
                        )
                result = crf_cache[crf_number]

                submitted_date_text = result.submitted_date if result else ""
                released_date_text = result.released_date if result else ""
                write_date_cell(worksheet[f"{columns.submitted}{row_idx}"], submitted_date_text)
                write_date_cell(worksheet[f"{columns.released}{row_idx}"], released_date_text)

                submitted_date = parse_sheet_date(submitted_date_text)
                released_date = parse_sheet_date(released_date_text)
                submitted_delta_end = submitted_date or today_sg
                released_delta_end = released_date or today_sg

                cell_updates = [
                    (
                        worksheet[f"{columns.submitted_delta}{row_idx}"],
                        calculate_delta_days(received_date, submitted_delta_end),
                    ),
                    (
                        worksheet[f"{columns.submitted_delta_excl_weekend}{row_idx}"],
                        calculate_delta_excluding_weekends(received_date, submitted_delta_end),
                    ),
                    (
                        worksheet[f"{columns.released_delta}{row_idx}"],
                        calculate_delta_days(received_date, released_delta_end),
                    ),
                    (
                        worksheet[f"{columns.released_delta_excl_weekend}{row_idx}"],
                        calculate_delta_excluding_weekends(received_date, released_delta_end),
                    ),
                ]
                for cell, value in cell_updates:
                    cell.value = value
                    clear_cell_style(cell)

                follow_up_reason = build_crf_follow_up_reason(
                    received_date=received_date,
                    crf_number=crf_number,
                    submitted_date=submitted_date,
                    released_date=released_date,
                    today_sg=today_sg,
                )
                if follow_up_reason and received_date:
                    reminder = CrfReminderItem(
                        row_number=row_idx,
                        crf_number=crf_number,
                        received_date=received_date.strftime("%d %b %Y"),
                        submitted_date=submitted_date.strftime("%d %b %Y") if submitted_date else "",
                        released_date=released_date.strftime("%d %b %Y") if released_date else "",
                        reason=follow_up_reason,
                    )
                    key = crf_reminder_key(reminder)
                    if not reminder_state.get(key):
                        reminders.append(reminder)
                        sent_reminder_keys.append(key)
                    else:
                        log_event(
                            f"CRF reminder already sent today for row {row_idx}: "
                            f"{crf_number or '(blank CRF number)'} - {follow_up_reason}",
                            SYNC_LOG_FILE,
                        )

                updated_rows += 1
                log_event(
                    f"Updated row {row_idx}: CRF={crf_number} "
                    f"Submitted={submitted_date_text} Released={released_date_text}",
                    SYNC_LOG_FILE,
                )

            refresh_delta_conditional_formatting(
                worksheet,
                start_row,
                worksheet.max_row,
                columns.submitted,
                columns.released,
                columns.submitted_delta,
                columns.released_delta,
                [
                    columns.submitted_delta,
                    columns.released_delta,
                    columns.submitted_delta_excl_weekend,
                    columns.released_delta_excl_weekend,
                ],
            )

            save_workbook_with_retry(
                workbook,
                workbook_path,
                LOCAL_WORKBOOK_RETRY_DELAY_MINUTES,
                LOCAL_WORKBOOK_SETTLE_SECONDS,
                SYNC_LOG_FILE,
            )
            break
        except Exception as exc:
            if not is_local_conflict_error(exc):
                log_exception("Unhandled CRF workbook update error.", exc, SYNC_LOG_FILE)
                raise
            log_exception("CRF workbook update conflict detected; will retry.", exc, SYNC_LOG_FILE)
            log_event(
                f"CRF workbook is still open or locked. Retrying in {LOCAL_WORKBOOK_RETRY_DELAY_MINUTES} minutes.",
                SYNC_LOG_FILE,
            )
            continue
        finally:
            if workbook is not None:
                workbook.close()

    reminder_sent = False
    if reminders:
        reminder_sent = send_crf_reminder_email(
            reminders,
            reminder_to,
            reminder_cc,
            reminder_bcc,
            preview_only=preview_reminder,
        )
    if reminder_sent and not preview_reminder:
        for key in sent_reminder_keys:
            reminder_state[key] = True
        save_crf_reminder_state(reminder_state)
    log_event(
        f"CRF workbook updated: {workbook_file}; rows updated={updated_rows}; "
        f"rows skipped (already released)={skipped_released_rows}; "
        f"rows skipped (blank/not applicable)={skipped_unusable_rows}",
        SYNC_LOG_FILE,
    )
    return updated_rows


def main() -> int:
    args = parse_args()
    class_identifiers = args.class_identifiers or DEFAULT_CLASS_CANDIDATES
    table_identifiers = args.table_identifiers or DEFAULT_TABLE_CANDIDATES
    crf_class_identifiers = args.crf_class_identifiers or class_identifiers
    crf_table_identifiers = args.crf_table_identifiers or table_identifiers

    if args.self_test or args.crf_self_test:
        exit_code = 0
        if args.self_test:
            exit_code |= run_self_test(
                workbook_file=args.workbook_file,
                worksheet_name=args.worksheet,
                reminder_to=args.reminder_to,
                reminder_cc=args.reminder_cc,
                reminder_bcc=args.reminder_bcc,
            )
        if args.crf_self_test:
            exit_code |= run_self_test(
                workbook_file=args.crf_workbook_file,
                worksheet_name=args.crf_worksheet,
                reminder_to=args.crf_reminder_to,
                reminder_cc=args.crf_reminder_cc,
                reminder_bcc=args.crf_reminder_bcc,
            )
        return exit_code

    if args.eco_number:
        agile_client = create_agile_client_from_file()
        result = fetch_eco_dates(
            client=agile_client,
            eco_number=args.eco_number,
            class_identifiers=class_identifiers,
            table_identifiers=table_identifiers,
            inspect=args.inspect,
        )

        print(f"ECO: {args.eco_number}")
        print(f"Class Identifier: {result.class_identifier}")
        print(f"Table Identifier: {result.table_identifier}")
        print(f"Submitted Date: {result.submitted_date or ''}")
        print(f"Released Date: {result.released_date or ''}")
        return 0

    if args.crf_number:
        agile_client = create_agile_client_from_file()
        result = fetch_eco_dates(
            client=agile_client,
            eco_number=args.crf_number,
            class_identifiers=crf_class_identifiers,
            table_identifiers=crf_table_identifiers,
            inspect=args.crf_inspect,
        )

        print(f"CRF: {args.crf_number}")
        print(f"Class Identifier: {result.class_identifier}")
        print(f"Table Identifier: {result.table_identifier}")
        print(f"Submitted Date: {result.submitted_date or ''}")
        print(f"Released Date: {result.released_date or ''}")
        return 0

    ran_workbook_update = False
    if args.workbook_file:
        update_workbook_dates(
            workbook_file=args.workbook_file,
            worksheet_name=args.worksheet,
            spec_award_column=args.spec_award_column,
            eco_column=args.eco_column,
            submitted_column=args.submitted_column,
            submitted_delta_column=args.submitted_delta_column,
            submitted_delta_excl_weekend_column=args.submitted_delta_excl_weekend_column,
            released_column=args.released_column,
            released_delta_column=args.released_delta_column,
            released_delta_excl_weekend_column=args.released_delta_excl_weekend_column,
            start_row=args.start_row,
            class_identifiers=class_identifiers,
            table_identifiers=table_identifiers,
            reminder_to=args.reminder_to,
            reminder_cc=args.reminder_cc,
            reminder_bcc=args.reminder_bcc,
            preview_reminder=args.preview_reminder,
        )
        ran_workbook_update = True

    if args.crf_workbook_file:
        update_crf_workbook_dates(
            workbook_file=args.crf_workbook_file,
            worksheet_name=args.crf_worksheet,
            crf_number_column=args.crf_number_column,
            received_date_column=args.crf_received_column,
            submitted_column=args.crf_submitted_column,
            submitted_delta_column=args.crf_submitted_delta_column,
            submitted_delta_excl_weekend_column=args.crf_submitted_delta_excl_weekend_column,
            released_column=args.crf_released_column,
            released_delta_column=args.crf_released_delta_column,
            released_delta_excl_weekend_column=args.crf_released_delta_excl_weekend_column,
            start_row=args.crf_start_row,
            class_identifiers=crf_class_identifiers,
            table_identifiers=crf_table_identifiers,
            reminder_to=args.crf_reminder_to,
            reminder_cc=args.crf_reminder_cc,
            reminder_bcc=args.crf_reminder_bcc,
            preview_reminder=args.crf_preview_reminder,
        )
        ran_workbook_update = True

    if not ran_workbook_update:
        raise RuntimeError(
            "Provide either an ECO number, --crf-number, --workbook-file, or --crf-workbook-file."
        )
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        raise SystemExit(1)
