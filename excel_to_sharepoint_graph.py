"""Append rows from a local Excel file into a SharePoint workbook table using Microsoft Graph.

This version uses delegated Microsoft Graph access with device-code sign-in and a local
MSAL token cache. It avoids username/password auth because Graph workbook APIs are
delegated-only and password-based flows are brittle under MFA and tenant policies.
"""

from __future__ import annotations

import argparse
import json
import os
import sys
import time
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple
from urllib.parse import quote

import msal
import requests
from openpyxl import load_workbook


GRAPH_BASE_URL = "https://graph.microsoft.com/v1.0"
GRAPH_SCOPES = ["Files.ReadWrite", "Sites.Read.All"]
DEFAULT_CACHE_FILE = ".msal_token_cache.bin"
DEFAULT_BATCH_SIZE = 50
SYSTEM_NUMBER_COLUMN = "System Number"


@dataclass
class Config:
    tenant_id: str
    client_id: str
    sharepoint_hostname: str
    sharepoint_site_path: str
    workbook_path: str
    table_name: str
    excel_file: str
    excel_sheet: Optional[str]
    field_map: Dict[str, str]
    batch_size: int
    cache_file: str
    dry_run: bool
    clear_source_on_success: bool


def parse_args() -> Config:
    parser = argparse.ArgumentParser(
        description="Append local Excel rows to a SharePoint workbook table via Microsoft Graph."
    )
    parser.add_argument("--tenant-id", default=os.getenv("SHAREPOINT_TENANT_ID"))
    parser.add_argument("--client-id", default=os.getenv("SHAREPOINT_CLIENT_ID"))
    parser.add_argument(
        "--sharepoint-hostname",
        default=os.getenv("SHAREPOINT_HOSTNAME"),
        help="Example: contoso.sharepoint.com",
    )
    parser.add_argument(
        "--sharepoint-site-path",
        default=os.getenv("SHAREPOINT_SITE_PATH"),
        help="Example: /sites/ECO-Tracker",
    )
    parser.add_argument(
        "--workbook-path",
        default=os.getenv("SHAREPOINT_WORKBOOK_PATH"),
        help="Example: Shared Documents/AMAT SGP ECO Tracker.xlsx",
    )
    parser.add_argument(
        "--table-name",
        default=os.getenv("TABLE_NAME"),
        help="Existing Excel table name in the SharePoint workbook",
    )
    parser.add_argument("--excel-file", default=os.getenv("EXCEL_FILE"))
    parser.add_argument("--excel-sheet", default=os.getenv("EXCEL_SHEET"))
    parser.add_argument(
        "--map",
        action="append",
        default=[],
        metavar="EXCEL_HEADER=TABLE_COLUMN",
        help="Map an Excel header to a workbook table column name. Can be used multiple times.",
    )
    parser.add_argument(
        "--batch-size",
        type=int,
        default=int(os.getenv("BATCH_SIZE", str(DEFAULT_BATCH_SIZE))),
        help="How many rows to send in one Graph request.",
    )
    parser.add_argument(
        "--cache-file",
        default=os.getenv("TOKEN_CACHE_FILE", DEFAULT_CACHE_FILE),
        help="Path to the local MSAL token cache file.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        default=os.getenv("DRY_RUN", "false").lower() in {"1", "true", "yes"},
        help="Print rows that would be uploaded without writing to SharePoint.",
    )
    parser.add_argument(
        "--clear-source-on-success",
        action="store_true",
        help="Clear the source workbook rows after a successful sync.",
    )

    args = parser.parse_args()
    validate_args(parser, args)
    field_map = parse_field_map(parser, args.map)

    return Config(
        tenant_id=args.tenant_id,
        client_id=args.client_id,
        sharepoint_hostname=args.sharepoint_hostname,
        sharepoint_site_path=args.sharepoint_site_path,
        workbook_path=args.workbook_path,
        table_name=args.table_name,
        excel_file=args.excel_file,
        excel_sheet=args.excel_sheet,
        field_map=field_map,
        batch_size=args.batch_size,
        cache_file=args.cache_file,
        dry_run=args.dry_run,
        clear_source_on_success=args.clear_source_on_success,
    )


def validate_args(parser: argparse.ArgumentParser, args: argparse.Namespace) -> None:
    missing = [
        name
        for name, value in [
            ("tenant_id", args.tenant_id),
            ("client_id", args.client_id),
            ("sharepoint_hostname", args.sharepoint_hostname),
            ("sharepoint_site_path", args.sharepoint_site_path),
            ("workbook_path", args.workbook_path),
            ("table_name", args.table_name),
            ("excel_file", args.excel_file),
        ]
        if not value
    ]

    if missing:
        parser.error("Missing required values: " + ", ".join(missing))
    if args.batch_size < 1:
        parser.error("--batch-size must be at least 1")


def parse_field_map(
    parser: argparse.ArgumentParser,
    mappings: List[str],
) -> Dict[str, str]:
    field_map: Dict[str, str] = {}
    for mapping in mappings:
        if "=" not in mapping:
            parser.error(f"Invalid --map value '{mapping}'. Use EXCEL_HEADER=TABLE_COLUMN.")
        excel_header, table_column = mapping.split("=", 1)
        field_map[excel_header.strip()] = table_column.strip()
    return field_map


def load_cache(path: str) -> msal.SerializableTokenCache:
    cache = msal.SerializableTokenCache()
    cache_path = Path(path)
    if cache_path.exists():
        cache.deserialize(cache_path.read_text(encoding="utf-8"))
    return cache


def save_cache(cache: msal.SerializableTokenCache, path: str) -> None:
    if not cache.has_state_changed:
        return
    Path(path).write_text(cache.serialize(), encoding="utf-8")


def acquire_token(config: Config) -> str:
    authority = f"https://login.microsoftonline.com/{config.tenant_id}"
    cache = load_cache(config.cache_file)
    app = msal.PublicClientApplication(
        client_id=config.client_id,
        authority=authority,
        token_cache=cache,
    )

    accounts = app.get_accounts()
    token_result = None
    if accounts:
        token_result = app.acquire_token_silent(GRAPH_SCOPES, account=accounts[0])

    if not token_result:
        flow = app.initiate_device_flow(scopes=GRAPH_SCOPES)
        if "user_code" not in flow:
            raise RuntimeError(f"Failed to start device login: {flow}")
        print(flow["message"])
        token_result = app.acquire_token_by_device_flow(flow)

    save_cache(cache, config.cache_file)

    if not token_result or "access_token" not in token_result:
        raise RuntimeError(
            "Could not acquire a Graph access token. "
            f"Details: {token_result.get('error_description') if token_result else 'unknown'}"
        )
    return token_result["access_token"]


def graph_request(
    method: str,
    url: str,
    token: str,
    *,
    params: Optional[Dict[str, Any]] = None,
    json_body: Optional[Dict[str, Any]] = None,
    retries: int = 4,
) -> requests.Response:
    headers = {"Authorization": f"Bearer {token}"}
    if json_body is not None:
        headers["Content-Type"] = "application/json"

    for attempt in range(retries):
        response = requests.request(
            method,
            url,
            headers=headers,
            params=params,
            json=json_body,
            timeout=60,
        )
        if response.status_code in {429, 503, 504} and attempt < retries - 1:
            retry_after = response.headers.get("Retry-After")
            sleep_seconds = int(retry_after) if retry_after and retry_after.isdigit() else 5
            print(
                f"Graph {method} {url} returned {response.status_code}. Retrying in {sleep_seconds} seconds..."
            )
            time.sleep(sleep_seconds)
            continue
        if response.status_code >= 400:
            raise requests.HTTPError(
                f"{method} {url} failed with status {response.status_code}: {response.text}",
                response=response,
            )
        return response

    raise RuntimeError(f"Graph request failed after {retries} attempts: {method} {url}")


def graph_get_json(
    url: str,
    token: str,
    *,
    params: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    return graph_request("GET", url, token, params=params).json()


def get_site_id(config: Config, token: str) -> str:
    url = f"{GRAPH_BASE_URL}/sites/{config.sharepoint_hostname}:{config.sharepoint_site_path}"
    payload = graph_get_json(url, token, params={"$select": "id,webUrl"})
    site_id = payload.get("id")
    if not site_id:
        raise RuntimeError("Could not resolve the SharePoint site id.")
    return site_id


def get_drive_item(config: Config, token: str, site_id: str) -> Tuple[str, str]:
    workbook_path = quote(config.workbook_path.strip("/"), safe="/")
    url = f"{GRAPH_BASE_URL}/sites/{site_id}/drive/root:/{workbook_path}"
    payload = graph_get_json(url, token, params={"$select": "id,webUrl,parentReference"})

    item_id = payload.get("id")
    drive_id = (payload.get("parentReference") or {}).get("driveId")
    if not item_id or not drive_id:
        raise RuntimeError("Could not resolve the workbook drive item.")
    return drive_id, item_id


def table_base_url(drive_id: str, item_id: str) -> str:
    return f"{GRAPH_BASE_URL}/drives/{drive_id}/items/{item_id}/workbook/tables"


def get_table_columns(drive_id: str, item_id: str, table_name: str, token: str) -> List[str]:
    url = f"{table_base_url(drive_id, item_id)}/{quote(table_name)}/columns"
    payload = graph_get_json(url, token, params={"$top": 999})
    columns = [col.get("name", "").strip() for col in payload.get("value", [])]
    columns = [col for col in columns if col]
    if not columns:
        raise RuntimeError(f"Could not read columns for table '{table_name}'.")
    return columns


def get_table_rows(
    drive_id: str,
    item_id: str,
    table_name: str,
    token: str,
) -> List[List[Any]]:
    rows: List[List[Any]] = []
    next_url = f"{table_base_url(drive_id, item_id)}/{quote(table_name)}/rows"
    params: Optional[Dict[str, Any]] = {"$top": 999}

    while next_url:
        payload = graph_get_json(next_url, token, params=params)
        for row in payload.get("value", []):
            values = row.get("values") or []
            if values and isinstance(values, list) and values and isinstance(values[0], list):
                rows.append(values[0])
            else:
                rows.append(values)
        next_link = payload.get("@odata.nextLink")
        next_url = next_link or ""
        params = None

    return rows


def read_existing_keys(
    drive_id: str,
    item_id: str,
    table_name: str,
    key_column: str,
    token: str,
) -> set[str]:
    columns = get_table_columns(drive_id, item_id, table_name, token)
    if key_column not in columns:
        return set()

    key_index = columns.index(key_column)
    existing_keys: set[str] = set()
    for row in get_table_rows(drive_id, item_id, table_name, token):
        if key_index < len(row) and row[key_index] is not None:
            existing_keys.add(str(row[key_index]).strip())
    return existing_keys


def batch_rows(items: List[List[Any]], batch_size: int) -> Iterable[List[List[Any]]]:
    for start in range(0, len(items), batch_size):
        yield items[start : start + batch_size]


def append_rows_to_graph_table(
    drive_id: str,
    item_id: str,
    table_name: str,
    token: str,
    rows: List[List[Any]],
    batch_size: int,
) -> int:
    written = 0
    url = f"{table_base_url(drive_id, item_id)}/{quote(table_name)}/rows"

    for chunk in batch_rows(rows, batch_size):
        graph_request("POST", url, token, json_body={"values": chunk})
        written += len(chunk)
        print(f"Appended {written}/{len(rows)} row(s).")

    return written


def load_rows_from_excel(path: str, sheet_name: Optional[str]) -> List[Dict[str, Any]]:
    workbook = load_workbook(path, data_only=True)
    try:
        sheet = workbook[sheet_name] if sheet_name else workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        data_rows: List[Dict[str, Any]] = []

        for row in rows[1:]:
            item: Dict[str, Any] = {}
            for header, value in zip(headers, row):
                if not header or value is None:
                    continue
                item[header] = value
            if item:
                data_rows.append(item)

        return data_rows
    finally:
        workbook.close()


def remove_excel_rows_by_keys(
    path: str,
    sheet_name: Optional[str],
    keys_to_remove: set[str],
    key_column_index: int = 1,
) -> None:
    if not keys_to_remove:
        return

    workbook = load_workbook(path)
    try:
        sheet = workbook[sheet_name] if sheet_name else workbook.active
        for row_idx in range(sheet.max_row, 1, -1):
            key_value = sheet.cell(row=row_idx, column=key_column_index).value
            if key_value is None:
                continue
            if str(key_value).strip() in keys_to_remove:
                sheet.delete_rows(row_idx, 1)
        workbook.save(path)
    finally:
        workbook.close()


def normalize_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.strftime("%d %b %Y")
    if isinstance(value, date):
        return value.strftime("%d %b %Y")
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float, str)):
        return value
    return str(value)


def apply_field_map(row: Dict[str, Any], field_map: Dict[str, str]) -> Dict[str, Any]:
    mapped: Dict[str, Any] = {}
    for excel_header, value in row.items():
        table_column = field_map.get(excel_header, excel_header)
        mapped[table_column] = normalize_value(value)
    return mapped


def align_row_to_table(row: Dict[str, Any], table_columns: List[str]) -> List[Any]:
    return [row.get(column) for column in table_columns]


def prepare_rows_for_table(
    source_rows: List[Dict[str, Any]],
    field_map: Dict[str, str],
    table_columns: List[str],
    dry_run: bool,
) -> List[List[Any]]:
    prepared_rows: List[List[Any]] = []
    for index, row in enumerate(source_rows, start=2):
        mapped = apply_field_map(row, field_map)
        aligned = align_row_to_table(mapped, table_columns)
        prepared_rows.append(aligned)
        if dry_run:
            print(f"[DRY RUN] Row {index}: {aligned}")
    return prepared_rows


def run_once(config: Config) -> int:
    source_rows = load_rows_from_excel(config.excel_file, config.excel_sheet)
    source_keys = {
        str(row.get(SYSTEM_NUMBER_COLUMN)).strip()
        for row in source_rows
        if row.get(SYSTEM_NUMBER_COLUMN) is not None and str(row.get(SYSTEM_NUMBER_COLUMN)).strip()
    }

    if not source_rows:
        print("No data rows found in the Excel file.")
        return 0

    token = acquire_token(config)
    site_id = get_site_id(config, token)
    drive_id, item_id = get_drive_item(config, token, site_id)
    table_columns = get_table_columns(drive_id, item_id, config.table_name, token)

    print(f"Found {len(source_rows)} source row(s).")
    print(f"Target workbook: {config.sharepoint_hostname}:{config.sharepoint_site_path}/{config.workbook_path}")
    print(f"Target table: {config.table_name}")
    print(f"Table columns: {', '.join(table_columns)}")

    prepared_rows = prepare_rows_for_table(
        source_rows=source_rows,
        field_map=config.field_map,
        table_columns=table_columns,
        dry_run=config.dry_run,
    )

    existing_keys = read_existing_keys(drive_id, item_id, config.table_name, SYSTEM_NUMBER_COLUMN, token)
    before_count = len(prepared_rows)
    if existing_keys:
        key_index = table_columns.index(SYSTEM_NUMBER_COLUMN) if SYSTEM_NUMBER_COLUMN in table_columns else None
        if key_index is not None:
            filtered_rows: List[List[Any]] = []
            seen_keys = set(existing_keys)
            for row in prepared_rows:
                key_value = ""
                if key_index < len(row) and row[key_index] is not None:
                    key_value = str(row[key_index]).strip()
                if key_value and key_value in seen_keys:
                    print(f"Skipping duplicate {SYSTEM_NUMBER_COLUMN}: {key_value}")
                    continue
                if key_value:
                    seen_keys.add(key_value)
                filtered_rows.append(row)
            prepared_rows = filtered_rows
    skipped_count = before_count - len(prepared_rows)
    if skipped_count:
        print(f"Skipped {skipped_count} duplicate row(s).")

    if config.dry_run:
        print("Dry run complete. No SharePoint data was changed.")
        return 0

    if not prepared_rows:
        print("No new rows to append after duplicate check.")
        if config.clear_source_on_success:
            remove_excel_rows_by_keys(config.excel_file, config.excel_sheet, source_keys)
            print("Processed rows removed from source workbook queue.")
        return 0

    written = append_rows_to_graph_table(
        drive_id=drive_id,
        item_id=item_id,
        table_name=config.table_name,
        token=token,
        rows=prepared_rows,
        batch_size=config.batch_size,
    )

    if config.clear_source_on_success:
        remove_excel_rows_by_keys(config.excel_file, config.excel_sheet, source_keys)
        print("Processed rows removed from source workbook queue.")

    print(f"Upload complete. {written} row(s) appended.")
    return 0


def main() -> int:
    config = parse_args()
    try:
        return run_once(config)
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        raise SystemExit(1)


if __name__ == "__main__":
    raise SystemExit(main())
