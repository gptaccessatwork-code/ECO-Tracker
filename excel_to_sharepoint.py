"""Append rows from a local Excel file into a workbook table.

Supported target modes:
- Microsoft Graph for a SharePoint-hosted workbook
- A locally synced OneDrive/SharePoint workbook file
"""

from __future__ import annotations

import argparse
import base64
import os
import sys
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional
from urllib.parse import quote

import msal
import requests
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter, range_boundaries
from workbook_sync_utils import (
    is_local_conflict_error,
    log_event,
    log_exception,
    save_workbook_with_retry,
    wait_for_workbook_ready,
)


GRAPH_BASE_URL = "https://graph.microsoft.com/v1.0"
GRAPH_SCOPES = ["Files.ReadWrite"]
DEFAULT_CACHE_FILE = ".msal_token_cache.bin"
DEFAULT_BATCH_SIZE = 50
DEFAULT_QUEUE_POLL_SECONDS = 10
DEFAULT_QUEUE_IDLE_CHECKS = 3
SYSTEM_NUMBER_COLUMN = "System Number"
LOCAL_WORKBOOK_RETRY_DELAY_MINUTES = 15
LOCAL_WORKBOOK_SETTLE_SECONDS = 30
SYNC_LOG_FILE = Path(__file__).with_name(".excel_to_sharepoint_sync.log")
SYNC_LOCK_FILE = Path(__file__).with_name(".excel_to_sharepoint_sync.lock")


@dataclass
class Config:
    tenant_id: Optional[str]
    client_id: Optional[str]
    workbook_url: Optional[str]
    sharepoint_hostname: Optional[str]
    sharepoint_site_path: Optional[str]
    workbook_path: Optional[str]
    target_workbook_file: Optional[str]
    table_name: str
    excel_file: str
    excel_sheet: Optional[str]
    field_map: Dict[str, str]
    batch_size: int
    cache_file: str
    dry_run: bool
    retry_delay_minutes: int
    clear_source_on_success: bool
    sync_until_empty: bool
    queue_poll_seconds: int
    queue_idle_checks: int


@dataclass(frozen=True)
class LocalTableContext:
    workbook: Any
    worksheet: Any
    table: Any


def parse_args() -> Config:
    parser = argparse.ArgumentParser(
        description="Append local Excel rows to a SharePoint workbook table."
    )
    parser.add_argument("--tenant-id", default=os.getenv("TENANT_ID"), required=False)
    parser.add_argument("--client-id", default=os.getenv("CLIENT_ID"), required=False)
    parser.add_argument(
        "--workbook-url",
        default=os.getenv("WORKBOOK_URL"),
        required=False,
        help="Full SharePoint/OneDrive workbook URL. Use this instead of hostname, site path, and workbook path.",
    )
    parser.add_argument(
        "--sharepoint-hostname",
        default=os.getenv("SHAREPOINT_HOSTNAME"),
        required=False,
        help="Example: contoso.sharepoint.com",
    )
    parser.add_argument(
        "--sharepoint-site-path",
        default=os.getenv("SHAREPOINT_SITE_PATH"),
        required=False,
        help="Example: /sites/ECO-Tracker",
    )
    parser.add_argument(
        "--workbook-path",
        default=os.getenv("WORKBOOK_PATH"),
        required=False,
        help="Path to the workbook inside the SharePoint document library, for example Shared Documents/Tracker.xlsx",
    )
    parser.add_argument(
        "--target-workbook-file",
        default=os.getenv("TARGET_WORKBOOK_FILE"),
        required=False,
        help="Local path to a synced OneDrive/SharePoint workbook file. Use this to avoid Graph authentication.",
    )
    parser.add_argument(
        "--table-name",
        default=os.getenv("TABLE_NAME"),
        required=False,
        help="Existing Excel table name in the SharePoint workbook",
    )
    parser.add_argument("--excel-file", default=os.getenv("EXCEL_FILE"), required=False)
    parser.add_argument("--excel-sheet", default=os.getenv("EXCEL_SHEET"), required=False)
    parser.add_argument(
        "--map",
        action="append",
        default=[],
        metavar="EXCEL_HEADER=TABLE_COLUMN",
        help="Map an Excel header to a workbook table column name. Can be used multiple times.",
    )
    parser.add_argument(
        "--batch-size",
        type=int,
        default=int(os.getenv("BATCH_SIZE", str(DEFAULT_BATCH_SIZE))),
        help="How many rows to send in one Graph request.",
    )
    parser.add_argument(
        "--cache-file",
        default=os.getenv("TOKEN_CACHE_FILE", DEFAULT_CACHE_FILE),
        help="Path to the local MSAL token cache file.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        default=os.getenv("DRY_RUN", "false").lower() in {"1", "true", "yes"},
        help="Print rows that would be uploaded without writing to SharePoint.",
    )
    parser.add_argument(
        "--retry-delay-minutes",
        type=int,
        default=int(os.getenv("RETRY_DELAY_MINUTES", str(LOCAL_WORKBOOK_RETRY_DELAY_MINUTES))),
        help="Minutes to wait between retries when the workbook is open or locked.",
    )
    parser.add_argument(
        "--clear-source-on-success",
        action="store_true",
        help="Clear the source workbook rows after a successful sync.",
    )
    parser.add_argument(
        "--sync-until-empty",
        action="store_true",
        help="Keep polling the source workbook until it stays empty for a few checks.",
    )
    parser.add_argument(
        "--queue-poll-seconds",
        type=int,
        default=int(os.getenv("QUEUE_POLL_SECONDS", str(DEFAULT_QUEUE_POLL_SECONDS))),
        help="Seconds to wait between empty-queue checks when --sync-until-empty is set.",
    )
    parser.add_argument(
        "--queue-idle-checks",
        type=int,
        default=int(os.getenv("QUEUE_IDLE_CHECKS", str(DEFAULT_QUEUE_IDLE_CHECKS))),
        help="How many consecutive empty-queue checks to tolerate before exiting.",
    )

    args = parser.parse_args()

    validate_args(parser, args)
    field_map = parse_field_map(parser, args.map)

    return Config(
        tenant_id=args.tenant_id,
        client_id=args.client_id,
        workbook_url=args.workbook_url,
        sharepoint_hostname=args.sharepoint_hostname,
        sharepoint_site_path=args.sharepoint_site_path,
        workbook_path=args.workbook_path,
        target_workbook_file=args.target_workbook_file,
        table_name=args.table_name,
        excel_file=args.excel_file,
        excel_sheet=args.excel_sheet,
        field_map=field_map,
        batch_size=args.batch_size,
        cache_file=args.cache_file,
        dry_run=args.dry_run,
        retry_delay_minutes=args.retry_delay_minutes,
        clear_source_on_success=args.clear_source_on_success,
        sync_until_empty=args.sync_until_empty,
        queue_poll_seconds=args.queue_poll_seconds,
        queue_idle_checks=args.queue_idle_checks,
    )


def validate_args(parser: argparse.ArgumentParser, args: argparse.Namespace) -> None:
    missing = [
        name
        for name, value in [
            ("table_name", args.table_name),
            ("excel_file", args.excel_file),
        ]
        if not value
    ]

    use_local_target = bool(args.target_workbook_file)
    if not use_local_target:
        for name, value in [
            ("tenant_id", args.tenant_id),
            ("client_id", args.client_id),
        ]:
            if not value:
                missing.append(name)

        if not args.workbook_url:
            missing.extend(
                name
                for name, value in [
                    ("sharepoint_hostname", args.sharepoint_hostname),
                    ("sharepoint_site_path", args.sharepoint_site_path),
                    ("workbook_path", args.workbook_path),
                ]
                if not value
            )

    if missing:
        parser.error("Missing required values: " + ", ".join(missing))

    if args.batch_size < 1:
        parser.error("--batch-size must be at least 1")
    if args.retry_delay_minutes < 1:
        parser.error("--retry-delay-minutes must be at least 1")
    if args.queue_poll_seconds < 1:
        parser.error("--queue-poll-seconds must be at least 1")
    if args.queue_idle_checks < 1:
        parser.error("--queue-idle-checks must be at least 1")


def parse_field_map(
    parser: argparse.ArgumentParser,
    mappings: List[str],
) -> Dict[str, str]:
    field_map: Dict[str, str] = {}
    for mapping in mappings:
        if "=" not in mapping:
            parser.error(f"Invalid --map value '{mapping}'. Use EXCEL_HEADER=TABLE_COLUMN.")
        excel_header, table_column = mapping.split("=", 1)
        field_map[excel_header.strip()] = table_column.strip()
    return field_map


def load_cache(path: str) -> msal.SerializableTokenCache:
    cache = msal.SerializableTokenCache()
    cache_path = Path(path)
    if cache_path.exists():
        cache.deserialize(cache_path.read_text(encoding="utf-8"))
    return cache


def save_cache(cache: msal.SerializableTokenCache, path: str) -> None:
    if not cache.has_state_changed:
        return
    Path(path).write_text(cache.serialize(), encoding="utf-8")


def get_access_token(config: Config) -> str:
    if not config.tenant_id or not config.client_id:
        raise ValueError("Tenant ID and client ID are required for Graph mode.")
    cache = load_cache(config.cache_file)
    app = msal.PublicClientApplication(
        client_id=config.client_id,
        authority=f"https://login.microsoftonline.com/{config.tenant_id}",
        token_cache=cache,
    )

    accounts = app.get_accounts()
    result = None
    if accounts:
        result = app.acquire_token_silent(GRAPH_SCOPES, account=accounts[0])

    if not result:
        flow = app.initiate_device_flow(scopes=GRAPH_SCOPES)
        if "user_code" not in flow:
            raise RuntimeError("Could not start device code flow.")
        print(flow["message"])
        result = app.acquire_token_by_device_flow(flow)

    save_cache(cache, config.cache_file)

    if not result or "access_token" not in result:
        error_text = "unknown error"
        if result:
            error_text = result.get("error_description") or result.get("error") or error_text
        raise RuntimeError("Could not acquire Graph access token: " + error_text)

    return result["access_token"]


def graph_request(method: str, url: str, token: str, **kwargs: Any) -> requests.Response:
    headers = kwargs.pop("headers", {})
    headers["Authorization"] = f"Bearer {token}"
    headers.setdefault("Content-Type", "application/json")
    response = requests.request(method, url, headers=headers, timeout=60, **kwargs)
    response.raise_for_status()
    return response


def normalize_site_path(path: str) -> str:
    path = path.strip()
    if not path.startswith("/"):
        path = "/" + path
    return path


def encode_workbook_path(path: str) -> str:
    return quote(path.strip("/"), safe="/")


def encode_sharing_url(url: str) -> str:
    encoded = base64.urlsafe_b64encode(url.encode("utf-8")).decode("utf-8")
    return "u!" + encoded.rstrip("=")


def resolve_site_id(config: Config, token: str) -> str:
    if not config.sharepoint_hostname or not config.sharepoint_site_path:
        raise ValueError("SharePoint hostname and site path are required without workbook URL.")
    site_path = normalize_site_path(config.sharepoint_site_path)
    url = f"{GRAPH_BASE_URL}/sites/{config.sharepoint_hostname}:{site_path}"
    response = graph_request("GET", url, token)
    return response.json()["id"]


def resolve_workbook_by_path_url(site_id: str, config: Config) -> str:
    if not config.workbook_path:
        raise ValueError("Workbook path is required without workbook URL.")
    workbook_path = encode_workbook_path(config.workbook_path)
    return (
        f"{GRAPH_BASE_URL}/sites/{site_id}/drive/root:/{workbook_path}:"
        f"/workbook"
    )


def resolve_workbook_by_sharing_url(config: Config, token: str) -> str:
    if not config.workbook_url:
        raise ValueError("Workbook URL is required.")

    share_id = encode_sharing_url(config.workbook_url)
    url = f"{GRAPH_BASE_URL}/shares/{share_id}/driveItem"
    response = graph_request("GET", url, token)
    drive_item = response.json()
    drive_id = drive_item.get("parentReference", {}).get("driveId")
    item_id = drive_item.get("id")

    if not drive_id or not item_id:
        raise RuntimeError("Could not resolve workbook drive and item IDs from the URL.")

    return f"{GRAPH_BASE_URL}/drives/{drive_id}/items/{item_id}/workbook"


def resolve_workbook_base_url(config: Config, token: str) -> str:
    if config.workbook_url:
        return resolve_workbook_by_sharing_url(config, token)

    site_id = resolve_site_id(config, token)
    return resolve_workbook_by_path_url(site_id, config)


def get_table_columns(workbook_base_url: str, table_name: str, token: str) -> List[str]:
    url = f"{workbook_base_url}/tables/{quote(table_name)}/columns?$select=name"
    response = graph_request("GET", url, token)
    payload = response.json()
    columns = [item["name"] for item in payload.get("value", []) if item.get("name")]
    if not columns:
        raise RuntimeError(
            f"Could not read columns for table '{table_name}'. Make sure the table exists."
        )
    return columns


def get_local_table(target_workbook_file: str, table_name: str) -> LocalTableContext:
    try:
        workbook = load_workbook(target_workbook_file)
    except PermissionError as exc:
        raise RuntimeError(
            f"Could not open local target workbook '{target_workbook_file}'. "
            "Close Excel or any app using the file, then try again."
        ) from exc
    for worksheet in workbook.worksheets:
        if table_name in worksheet.tables:
            return LocalTableContext(workbook, worksheet, worksheet.tables[table_name])
        for table in worksheet.tables.values():
            if getattr(table, "displayName", None) == table_name:
                return LocalTableContext(workbook, worksheet, table)
    workbook.close()
    raise RuntimeError(
        f"Could not find table '{table_name}' in local workbook '{target_workbook_file}'."
    )


def get_local_table_columns(worksheet, table) -> List[str]:
    min_col, min_row, max_col, _ = range_boundaries(table.ref)
    headers: List[str] = []
    for col_idx in range(min_col, max_col + 1):
        value = worksheet.cell(row=min_row, column=col_idx).value
        headers.append("" if value is None else str(value).strip())
    return headers


def get_local_table_rows(worksheet, table) -> List[List[Any]]:
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    rows: List[List[Any]] = []
    for row_idx in range(min_row + 1, max_row + 1):
        row_values: List[Any] = []
        for col_idx in range(min_col, max_col + 1):
            row_values.append(worksheet.cell(row=row_idx, column=col_idx).value)
        rows.append(row_values)
    return rows


def is_table_row_empty(row_values: List[Any]) -> bool:
    for value in row_values:
        if value is None:
            continue
        if isinstance(value, str) and not value.strip():
            continue
        return False
    return True


def find_last_populated_table_row(worksheet, table) -> int:
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    last_populated_row = min_row

    for row_idx in range(min_row + 1, max_row + 1):
        row_values = [
            worksheet.cell(row=row_idx, column=col_idx).value
            for col_idx in range(min_col, max_col + 1)
        ]
        if not is_table_row_empty(row_values):
            last_populated_row = row_idx

    return last_populated_row


def read_local_table_columns(target_workbook_file: str, table_name: str) -> List[str]:
    context = get_local_table(target_workbook_file, table_name)
    try:
        return get_local_table_columns(context.worksheet, context.table)
    finally:
        context.workbook.close()


def read_local_existing_keys(
    target_workbook_file: str,
    table_name: str,
    key_column: str,
) -> set[str]:
    context = get_local_table(target_workbook_file, table_name)
    try:
        columns = get_local_table_columns(context.worksheet, context.table)
        if key_column not in columns:
            return set()
        key_index = columns.index(key_column)
        existing_keys: set[str] = set()
        for row in get_local_table_rows(context.worksheet, context.table):
            if key_index < len(row) and row[key_index] is not None:
                existing_keys.add(str(row[key_index]).strip())
        return existing_keys
    finally:
        context.workbook.close()


def append_rows_to_local_table(
    target_workbook_file: str,
    table_name: str,
    rows: List[List[Any]],
) -> None:
    log_event(f"Opening target workbook: {target_workbook_file}", SYNC_LOG_FILE)
    context = get_local_table(target_workbook_file, table_name)
    try:
        if getattr(context.table, "totalsRowShown", False):
            raise RuntimeError(
                f"Table '{table_name}' uses a totals row. This script currently expects a normal table without totals."
            )

        min_col, min_row, max_col, max_row = range_boundaries(context.table.ref)
        last_populated_row = find_last_populated_table_row(context.worksheet, context.table)
        next_row = last_populated_row + 1

        log_event(
            f"Writing {len(rows)} row(s) into table '{table_name}'...",
            SYNC_LOG_FILE,
        )
        for row_values in rows:
            for offset, value in enumerate(row_values):
                context.worksheet.cell(row=next_row, column=min_col + offset, value=value)
            next_row += 1

        new_max_row = max(max_row, last_populated_row + len(rows))
        context.table.ref = (
            f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{new_max_row}"
        )
        save_workbook_with_retry(
            context.workbook,
            target_workbook_file,
            LOCAL_WORKBOOK_RETRY_DELAY_MINUTES,
            LOCAL_WORKBOOK_SETTLE_SECONDS,
            SYNC_LOG_FILE,
        )
    finally:
        context.workbook.close()


def write_rows_to_local_workbook_with_retry(
    config: Config,
) -> int:
    if not config.target_workbook_file:
        raise ValueError("A local target workbook file is required.")

    while True:
        try:
            wait_for_workbook_ready(
                config.target_workbook_file,
                config.retry_delay_minutes,
                LOCAL_WORKBOOK_SETTLE_SECONDS,
                SYNC_LOG_FILE,
            )
            if config.excel_file and Path(config.excel_file).exists():
                wait_for_workbook_ready(
                    config.excel_file,
                    config.retry_delay_minutes,
                    LOCAL_WORKBOOK_SETTLE_SECONDS,
                    SYNC_LOG_FILE,
                )

            source_rows = load_rows_from_excel(config.excel_file, config.excel_sheet)
            if not source_rows:
                log_event("No source rows found in the staging workbook.", SYNC_LOG_FILE)
                return 0

            source_keys = {
                str(row.get(SYSTEM_NUMBER_COLUMN)).strip()
                for row in source_rows
                if row.get(SYSTEM_NUMBER_COLUMN) is not None
                and str(row.get(SYSTEM_NUMBER_COLUMN)).strip()
            }

            table_columns = read_local_table_columns(
                config.target_workbook_file, config.table_name
            )
            prepared_rows = prepare_rows_for_table(
                source_rows=source_rows,
                field_map=config.field_map,
                table_columns=table_columns,
                dry_run=False,
            )
            existing_keys = read_local_existing_keys(
                config.target_workbook_file, config.table_name, SYSTEM_NUMBER_COLUMN
            )
            rows_to_write = filter_duplicate_rows(
                prepared_rows, table_columns, existing_keys, SYSTEM_NUMBER_COLUMN
            )

            if not rows_to_write:
                log_event("No new rows to append after duplicate check.", SYNC_LOG_FILE)
                if config.clear_source_on_success:
                    remove_excel_rows_by_keys(
                        config.excel_file, config.excel_sheet, source_keys
                    )
                    log_event(
                        "Processed rows removed from source workbook queue.",
                        SYNC_LOG_FILE,
                    )
                return 0

            append_rows_to_local_table(
                config.target_workbook_file, config.table_name, rows_to_write
            )
            if config.clear_source_on_success:
                remove_excel_rows_by_keys(config.excel_file, config.excel_sheet, source_keys)
                log_event("Processed rows removed from source workbook queue.", SYNC_LOG_FILE)
            return len(rows_to_write)
        except Exception as exc:
            if not is_local_conflict_error(exc):
                log_exception("Non-lock error while writing to local workbook.", exc)
                raise
            log_exception("Workbook save conflict detected; will retry.", exc)
            log_event(
                f"Workbook is still open or locked. Retrying in {config.retry_delay_minutes} minutes.",
                SYNC_LOG_FILE,
            )
            continue


def load_rows_from_excel(path: str, sheet_name: Optional[str]) -> List[Dict[str, Any]]:
    workbook = load_workbook(path, data_only=True)
    try:
        sheet = workbook[sheet_name] if sheet_name else workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        data_rows: List[Dict[str, Any]] = []

        for row in rows[1:]:
            item: Dict[str, Any] = {}
            for header, value in zip(headers, row):
                if not header or value is None:
                    continue
                item[header] = value
            if item:
                data_rows.append(item)

        return data_rows
    finally:
        workbook.close()


def remove_excel_rows_by_keys(
    path: str,
    sheet_name: Optional[str],
    keys_to_remove: set[str],
    key_column_index: int = 1,
    log_file: Path = SYNC_LOG_FILE,
) -> None:
    if not keys_to_remove:
        return

    workbook = load_workbook(path)
    try:
        sheet = workbook[sheet_name] if sheet_name else workbook.active
        for row_idx in range(sheet.max_row, 1, -1):
            key_value = sheet.cell(row=row_idx, column=key_column_index).value
            if key_value is None:
                continue
            if str(key_value).strip() in keys_to_remove:
                sheet.delete_rows(row_idx, 1)
        save_workbook_with_retry(
            workbook,
            path,
            LOCAL_WORKBOOK_RETRY_DELAY_MINUTES,
            LOCAL_WORKBOOK_SETTLE_SECONDS,
            log_file,
        )
    finally:
        workbook.close()


def normalize_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.strftime("%d %b %Y")
    if isinstance(value, date):
        return value.strftime("%d %b %Y")
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float, str)):
        return value
    return str(value)


def apply_field_map(row: Dict[str, Any], field_map: Dict[str, str]) -> Dict[str, Any]:
    mapped: Dict[str, Any] = {}
    for excel_header, value in row.items():
        table_column = field_map.get(excel_header, excel_header)
        mapped[table_column] = normalize_value(value)
    return mapped


def align_row_to_table(row: Dict[str, Any], table_columns: List[str]) -> List[Any]:
    return [row.get(column) for column in table_columns]


def prepare_rows_for_table(
    source_rows: List[Dict[str, Any]],
    field_map: Dict[str, str],
    table_columns: List[str],
    dry_run: bool,
) -> List[List[Any]]:
    prepared_rows: List[List[Any]] = []
    for index, row in enumerate(source_rows, start=2):
        mapped = apply_field_map(row, field_map)
        aligned = align_row_to_table(mapped, table_columns)
        prepared_rows.append(aligned)
        if dry_run:
            print(f"[DRY RUN] Row {index}: {aligned}")
    return prepared_rows


def resolve_table_columns(config: Config) -> tuple[Optional[str], Optional[str], List[str]]:
    if config.target_workbook_file:
        return (
            None,
            None,
            read_local_table_columns(config.target_workbook_file, config.table_name),
        )

    token = get_access_token(config)
    workbook_base_url = resolve_workbook_base_url(config, token)
    table_columns = get_table_columns(workbook_base_url, config.table_name, token)
    return workbook_base_url, token, table_columns


def filter_duplicate_rows(
    prepared_rows: List[List[Any]],
    table_columns: List[str],
    existing_keys: set[str],
    key_column: str,
) -> List[List[Any]]:
    if key_column not in table_columns:
        return prepared_rows

    key_index = table_columns.index(key_column)
    filtered_rows: List[List[Any]] = []
    seen_keys = set(existing_keys)

    for row in prepared_rows:
        key_value = ""
        if key_index < len(row) and row[key_index] is not None:
            key_value = str(row[key_index]).strip()

        if key_value and key_value in seen_keys:
            print(f"Skipping duplicate {key_column}: {key_value}")
            continue

        if key_value:
            seen_keys.add(key_value)
        filtered_rows.append(row)

    return filtered_rows


def batch_rows(items: List[List[Any]], batch_size: int) -> Iterable[List[List[Any]]]:
    for start in range(0, len(items), batch_size):
        yield items[start : start + batch_size]


def acquire_sync_lock(lock_file: Path, log_file: Path) -> Optional[Any]:
    try:
        handle = lock_file.open("x", encoding="utf-8")
    except FileExistsError:
        log_event("A Spec Award sync worker is already running; skipping this launch.", log_file)
        return None

    handle.write(f"pid={os.getpid()}\nstarted={datetime.now().isoformat()}\n")
    handle.flush()
    return handle


def release_sync_lock(handle: Any, lock_file: Path) -> None:
    try:
        handle.close()
    finally:
        try:
            lock_file.unlink()
        except FileNotFoundError:
            pass


def sync_until_queue_is_empty(config: Config) -> int:
    total_written = 0
    idle_checks = 0

    while True:
        source_rows = load_rows_from_excel(config.excel_file, config.excel_sheet)
        if not source_rows:
            idle_checks += 1
            if idle_checks >= config.queue_idle_checks:
                return total_written
            log_event(
                f"Spec Award queue is empty. Rechecking in {config.queue_poll_seconds} seconds...",
                SYNC_LOG_FILE,
            )
            time.sleep(config.queue_poll_seconds)
            continue

        idle_checks = 0
        written = write_rows_to_local_workbook_with_retry(config=config)
        total_written += written

        if written > 0:
            log_event(
                f"Spec Award sync pass complete. Total appended so far: {total_written} row(s).",
                SYNC_LOG_FILE,
            )

    return total_written


def append_rows_to_table(
    workbook_base_url: str,
    table_name: str,
    token: str,
    rows: List[List[Any]],
) -> None:
    url = f"{workbook_base_url}/tables/{quote(table_name)}/rows/add"
    payload = {"values": rows}
    graph_request("POST", url, token, json=payload)


def run_once(config: Config) -> int:
    if config.target_workbook_file:
        if config.dry_run:
            source_rows = load_rows_from_excel(config.excel_file, config.excel_sheet)
            if not source_rows:
                print("No data rows found in the Excel file.")
                return 0

            table_columns = read_local_table_columns(config.target_workbook_file, config.table_name)
            print(f"Found {len(source_rows)} source row(s).")
            print(f"Target workbook: {config.target_workbook_file}")
            print(f"Target table: {config.table_name}")
            print(f"Table columns: {', '.join(table_columns)}")

            prepare_rows_for_table(
                source_rows=source_rows,
                field_map=config.field_map,
                table_columns=table_columns,
                dry_run=True,
            )
            print("Dry run complete. No SharePoint data was changed.")
            return 0

        written = write_rows_to_local_workbook_with_retry(config=config)
        print(f"Appended {written} row(s).")
    else:
        source_rows = load_rows_from_excel(config.excel_file, config.excel_sheet)
        source_keys = {
            str(row.get(SYSTEM_NUMBER_COLUMN)).strip()
            for row in source_rows
            if row.get(SYSTEM_NUMBER_COLUMN) is not None
            and str(row.get(SYSTEM_NUMBER_COLUMN)).strip()
        }

        if not source_rows:
            print("No data rows found in the Excel file.")
            return 0

        workbook_base_url, token, table_columns = resolve_table_columns(config)

        print(f"Found {len(source_rows)} source row(s).")
        print(
            f"Target workbook: "
            f"{config.target_workbook_file or config.workbook_url or config.workbook_path}"
        )
        print(f"Target table: {config.table_name}")
        print(f"Table columns: {', '.join(table_columns)}")

        prepared_rows = prepare_rows_for_table(
            source_rows=source_rows,
            field_map=config.field_map,
            table_columns=table_columns,
            dry_run=config.dry_run,
        )

        if config.dry_run:
            print("Dry run complete. No SharePoint data was changed.")
            return 0

        if not prepared_rows:
            print("No new rows to append after duplicate check.")
            if config.clear_source_on_success:
                remove_excel_rows_by_keys(config.excel_file, config.excel_sheet, source_keys)
                print("Processed rows removed from source workbook queue.")
            return 0

        written = 0
        for chunk in batch_rows(prepared_rows, config.batch_size):
            append_rows_to_table(workbook_base_url, config.table_name, token, chunk)
            written += len(chunk)
            print(f"Appended {written}/{len(prepared_rows)} row(s).")

    log_event(f"Upload complete. {written} row(s) appended.", SYNC_LOG_FILE)
    return 0


def main() -> int:
    config = parse_args()
    if config.sync_until_empty and config.target_workbook_file:
        lock_handle = acquire_sync_lock(SYNC_LOCK_FILE, SYNC_LOG_FILE)
        if lock_handle is None:
            return 0
        try:
            return sync_until_queue_is_empty(config)
        finally:
            release_sync_lock(lock_handle, SYNC_LOCK_FILE)
    return run_once(config)


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except requests.HTTPError as exc:
        log_exception("Graph request failed.", exc, SYNC_LOG_FILE)
        print(f"Graph request failed: {exc}", file=sys.stderr)
        if exc.response is not None:
            print(exc.response.text, file=sys.stderr)
        raise SystemExit(1)
    except Exception as exc:
        log_exception("Unhandled error.", exc, SYNC_LOG_FILE)
        print(f"Error: {exc}", file=sys.stderr)
        raise SystemExit(1)
